from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DB_EXPORT_DIR = BASE_DIR / "db_export"
OUTPUT_XLSX = BASE_DIR / "org_dashboard_google_sheet_template.xlsx"

SHEETS = [
    ("guide", None),
    ("sections", DB_EXPORT_DIR / "sections.csv"),
    ("organizations", DB_EXPORT_DIR / "organizations.csv"),
    ("people", DB_EXPORT_DIR / "people.csv"),
    ("assignments", DB_EXPORT_DIR / "assignments.csv"),
    ("role_rules", DB_EXPORT_DIR / "role_rules.csv"),
]

GUIDE_ROWS = [
    ["조직 대시보드 구글시트 운영 가이드", ""],
    ["목적", "정적 대시보드용 조직 데이터를 온라인에서 수정 가능한 형태로 운영하기 위한 원본 시트입니다."],
    ["권장 사용법", "이 파일을 구글시트로 업로드한 뒤 각 탭을 직접 수정합니다."],
    ["수정 우선순위", "1) assignments 2) organizations 3) sections 4) role_rules"],
    ["핵심 탭", "assignments"],
    ["assignments 설명", "사람-조직 배치 마스터. 인사 이동이나 직책 수정은 대부분 여기서 처리합니다."],
    ["organizations 설명", "조직 트리와 조직 코드 관리. 그룹/센터/TF/파트/팀 구조 수정 시 사용합니다."],
    ["sections 설명", "상위 부 단위 이름과 총괄자 직함 관리용입니다."],
    ["role_rules 설명", "대시보드 예외 규칙 관리용입니다. SS&C TF 비카운트 명단 같은 규칙을 둡니다."],
    ["주의사항", "person_id, org_id, assignment_id는 가능하면 유지하세요. 이름만 바꾸기보다 기존 ID를 유지하는 편이 안전합니다."],
    ["업로드 방법", "구글 드라이브 > 새로 만들기 > 파일 업로드 > 이 xlsx 업로드 > 구글시트로 열기"],
    ["다음 단계", "운영이 안정되면 대시보드가 이 구글시트를 직접 읽도록 연결할 수 있습니다."],
]


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        return list(csv.reader(fp))


def autofit_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value) + 2, 40))

    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def style_header(ws, header_row: int = 1) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font


def populate_guide(ws) -> None:
    for row in GUIDE_ROWS:
        ws.append(row)
    ws["A1"].font = Font(size=14, bold=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110
    for cell in ws["A"]:
        cell.font = Font(bold=cell.row == 1)


def populate_csv_sheet(ws, csv_path: Path) -> None:
    for row in read_csv(csv_path):
        ws.append(row)
    style_header(ws, 1)
    ws.freeze_panes = "A2"
    autofit_columns(ws)


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, csv_path in SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        if csv_path is None:
            populate_guide(ws)
        else:
            populate_csv_sheet(ws, csv_path)

    wb.save(OUTPUT_XLSX)
    print(f"Workbook created: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
