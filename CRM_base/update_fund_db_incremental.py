import argparse
import hashlib
import json
import math
import re
import struct
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT_DIR / "DB sources"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

TABLES = (
    "funds",
    "fund_assets",
    "asset_master",
    "asset_fund_links",
    "lender_exposures",
    "beneficiary_exposures",
    "aum_snapshots",
    "fund_lifecycle",
)

FUND_ID_RE = re.compile(r"^[A-Za-z0-9_-]+$")


def clean_text(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\xa0", " ").strip()
    if text.lower() in {"", "nan", "none", "nat", "null", "undefined"}:
        return None
    return text


def clean_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def clean_number(value, integer=False):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        text = value.replace(",", "").replace("%", "").strip()
        if not text or text.lower() in {"nan", "none", "null", "-"}:
            return None
        value = text
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if integer or number.is_integer():
        return int(round(number))
    return number


def clean_postgres_real(value):
    number = clean_number(value)
    if number is None:
        return None
    float32 = struct.unpack("!f", struct.pack("!f", float(number)))[0]
    return float(format(float32, ".6g"))


def yn_bool(value):
    text = clean_text(value)
    if text is None:
        return None
    if text.upper() in {"Y", "YES", "TRUE", "1"}:
        return True
    if text.upper() in {"N", "NO", "FALSE", "0"}:
        return False
    return None


def valid_fund_id(value):
    text = clean_text(value)
    return bool(text and FUND_ID_RE.fullmatch(text))


def make_id(prefix, *parts):
    raw = "|".join(clean_text(part) or "" for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}_{digest}"


def make_asset_id(asset_code):
    digest = hashlib.sha1(f"asset_code:{asset_code}".encode("utf-8")).hexdigest()[:12]
    return f"ast_{digest}"


def read_env():
    env_path = ROOT_DIR / ".env"
    values = {}
    supabase_keys = []
    for raw_line in env_path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key == "SUPABASE_KEY":
            supabase_keys.append(value)
        else:
            values[key] = value
    service_key = next((key for key in supabase_keys if key.startswith("sb_secret_")), None)
    fallback_key = next((key for key in supabase_keys if key.startswith("sb_publishable_")), None)
    api_key = service_key or fallback_key
    if not values.get("SUPABASE_URL") or not api_key:
        raise RuntimeError("SUPABASE_URL and a Supabase API key are required in .env")
    return values["SUPABASE_URL"].rstrip("/"), api_key


class PostgrestClient:
    def __init__(self, url, api_key):
        self.base_url = f"{url}/rest/v1"
        self.headers = {
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }

    def request(self, method, table, params=None, payload=None, prefer=None, extra_headers=None):
        query = f"?{urlencode(params or {}, doseq=True, safe=',.*()')}" if params else ""
        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer
        if extra_headers:
            headers.update(extra_headers)
        body = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = Request(f"{self.base_url}/{table}{query}", data=body, headers=headers, method=method)
        try:
            with urlopen(req, timeout=120) as response:
                raw = response.read()
                return json.loads(raw.decode("utf-8")) if raw else None
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"PostgREST {method} {table} failed ({exc.code}): {detail}") from exc

    def fetch_all(self, table, select="*"):
        records = []
        page_size = 1000
        start = 0
        while True:
            rows = self.request(
                "GET",
                table,
                params={"select": select},
                extra_headers={"Range": f"{start}-{start + page_size - 1}", "Range-Unit": "items"},
            ) or []
            records.extend(rows)
            if len(rows) < page_size:
                break
            start += page_size
        return records

    def write_grouped(self, table, records, on_conflict=None, batch_size=150):
        if not records:
            return 0
        groups = defaultdict(list)
        for record in records:
            groups[tuple(sorted(record))].append(record)
        total = 0
        for group in groups.values():
            for start in range(0, len(group), batch_size):
                chunk = group[start : start + batch_size]
                params = {"on_conflict": on_conflict} if on_conflict else None
                prefer = "return=minimal"
                if on_conflict:
                    prefer = "resolution=merge-duplicates,return=minimal"
                self.request("POST", table, params=params, payload=chunk, prefer=prefer)
                total += len(chunk)
        return total


def read_sheet(path, header_row=1):
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    headers = [clean_text(worksheet.cell(header_row, col).value) for col in range(1, worksheet.max_column + 1)]
    rows = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, col).value for col in range(1, worksheet.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({header: values[index] for index, header in enumerate(headers) if header})
    workbook.close()
    return rows


def read_aum_sheet(path):
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    rows = []
    for row_number in range(3, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, col).value for col in range(1, 42)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            {
                "fund_id": values[0],
                "short_name": values[1],
                "fund_name": values[2],
                "asset_type": values[3],
                "asset_name": values[4],
                "status": values[5],
                "incorporation_date": values[6],
                "setup_date": values[7],
                "maturity_date": values[8],
                "base_date": values[9],
                "base_price": values[10],
                "net_asset_value": values[11],
                "input_date": values[12],
                "equity": values[13],
                "loan": values[14],
                "deposit": values[15],
                "aum": values[16],
                "invested_equity": values[17],
                "invested_loan": values[18],
                "invested_deposit": values[19],
                "invested_aum": values[20],
                "beneficiaries": values[21],
                "lenders": values[22],
                "automation": values[23],
                "dept": values[24],
                "manager": values[25],
                "location": values[31],
                "fund_class": values[32],
                "is_delegated": values[33],
                "fund_type": values[34],
                "strategy": values[35],
                "parent_child": values[36],
                "multi_class": values[37],
                "fund_shape": values[38],
                "recruitment": values[39],
                "fund_structure": values[40],
            }
        )
    workbook.close()
    return rows


def unique_index(rows, key_func):
    grouped = defaultdict(list)
    for row in rows:
        key = key_func(row)
        if key is None or (isinstance(key, tuple) and any(part is None for part in key)):
            continue
        grouped[key].append(row)
    unique = {key: values[0] for key, values in grouped.items() if len(values) == 1}
    collisions = {str(key): len(values) for key, values in grouped.items() if len(values) > 1}
    return unique, collisions


def distinct_join(values):
    result = []
    for value in values:
        text = clean_text(value)
        if text and text not in result:
            result.append(text)
    return " / ".join(result) if result else None


def distinct_summary(values):
    result = []
    for value in values:
        text = clean_text(value)
        if text and text not in result:
            result.append(text)
    return ", ".join(result) if result else None


def sum_or_none(values):
    numbers = [clean_number(value, integer=True) for value in values]
    numbers = [number for number in numbers if number is not None]
    return sum(numbers) if numbers else None


def first_non_empty(values, cleaner=clean_text):
    for value in values:
        cleaned = cleaner(value)
        if cleaned is not None:
            return cleaned
    return None


def same_value(left, right):
    if left is None and right is None:
        return True
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-9, abs_tol=1e-6)
    return left == right


def changed_fields(existing, expected):
    return [key for key, value in expected.items() if key not in {"id"} and not same_value(existing.get(key), value)]


def merge_metadata(existing, updates):
    base = existing.get("metadata") if existing else None
    metadata = dict(base) if isinstance(base, dict) else {}
    metadata.update(updates)
    return metadata


def source_files(source_date):
    expected = {
        "fund": SOURCE_DIR / f"펀드 관리_{source_date}.xlsx",
        "aum": SOURCE_DIR / f"펀드 AUM 관리_{source_date}.xlsx",
        "asset_manage": SOURCE_DIR / f"투자 자산 관리_{source_date}.xlsx",
        "fund_asset": SOURCE_DIR / f"투자 자산 조회_{source_date}.xlsx",
        "lender": SOURCE_DIR / f"대주 정보 조회_{source_date}.xlsx",
        "beneficiary": SOURCE_DIR / f"수익자 정보 조회_{source_date}.xlsx",
    }
    missing = [str(path) for path in expected.values() if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing source workbooks: {missing}")
    return expected


def load_sources(paths):
    return {
        "fund": read_sheet(paths["fund"]),
        "aum": read_aum_sheet(paths["aum"]),
        "asset_manage": read_sheet(paths["asset_manage"]),
        "fund_asset": read_sheet(paths["fund_asset"]),
        "lender": read_sheet(paths["lender"]),
        "beneficiary": read_sheet(paths["beneficiary"]),
    }


def build_asset_classification(rows):
    grouped = defaultdict(list)
    for row in rows:
        fund_id = clean_text(row.get("펀드코드"))
        if fund_id:
            grouped[fund_id].append(row)
    result = {}
    for fund_id, values in grouped.items():
        result[fund_id] = {
            "base": distinct_summary(row.get("기초자산") for row in values),
            "nature": distinct_summary(row.get("자산성격") for row in values),
            "stage": distinct_summary(row.get("사업단계") for row in values),
        }
    return result


def build_fund_records(sources, paths):
    aum_rows = {}
    for row in sources["aum"]:
        fund_id = clean_text(row.get("fund_id"))
        if valid_fund_id(fund_id) and clean_text(row.get("fund_name")):
            if fund_id in aum_rows:
                raise RuntimeError(f"Duplicate fund_id in AUM workbook: {fund_id}")
            aum_rows[fund_id] = row

    classifications = build_asset_classification(sources["fund_asset"])
    records = []
    lifecycle = []
    snapshots = []
    seen = set()
    for row in sources["fund"]:
        fund_id = clean_text(row.get("펀드코드"))
        fund_name = clean_text(row.get("펀드명"))
        if not valid_fund_id(fund_id) or not fund_name:
            continue
        if fund_id in seen:
            raise RuntimeError(f"Duplicate fund_id in fund workbook: {fund_id}")
        seen.add(fund_id)
        classification = classifications.get(fund_id, {})
        record = {
            "fund_id": fund_id,
            "short_name": clean_text(row.get("약칭")),
            "fund_name": fund_name,
            "sector": clean_text(row.get("투자섹터")),
            "asset_name": clean_text(row.get("자산명")),
            "status": clean_text(row.get("운용상태")),
            "location": clean_text(row.get("국내/해외")),
            "setup_date": clean_date(row.get("최초 설정일")),
            "maturity_date": clean_date(row.get("만기일")),
            "termination_date": clean_date(row.get("해지일")),
            "dept": clean_text(row.get("부서(운용)")),
            "manager": clean_text(row.get("담당자(운용)")),
            "project_mission_name": clean_text(row.get("자산명")),
            "notion_base_asset_class": classification.get("base") or clean_text(row.get("투자섹터")),
            "notion_asset_nature_class": classification.get("nature") or clean_text(row.get("펀드유형")),
            "notion_holding_type_class": clean_text(row.get("모자구분")),
            "notion_business_stage_class": classification.get("stage"),
            "notion_investment_strategy_class": clean_text(row.get("투자전략")),
            "notion_vehicle_class": clean_text(row.get("Vehicle구분")),
            "recruitment_type": clean_text(row.get("모집형태")),
            "legal_form": clean_text(row.get("법적형태")),
            "fund_class": clean_text(row.get("펀드분류")),
            "fund_type": clean_text(row.get("펀드유형")),
            "division": clean_text(row.get("담당부문(운용)")),
            "primary_region": clean_text(row.get("주요투자지역")),
            "is_development": clean_text(row.get("개발여부")),
            "is_delegated": clean_text(row.get("위탁운용여부")),
        }
        aum = aum_rows.get(fund_id)
        if aum:
            record.update(
                {
                    "aum_base_date": clean_date(aum.get("base_date")),
                    "base_price": clean_number(aum.get("base_price")),
                    "net_asset_value": clean_number(aum.get("net_asset_value"), integer=True),
                    "aum_input_date": clean_date(aum.get("input_date")),
                    "equity_won": clean_number(aum.get("equity"), integer=True),
                    "loan_won": clean_number(aum.get("loan"), integer=True),
                    "deposit_won": clean_number(aum.get("deposit"), integer=True),
                    "benchmark_aum": clean_number(aum.get("aum"), integer=True),
                    "invested_equity_won": clean_number(aum.get("invested_equity"), integer=True),
                    "invested_loan_won": clean_number(aum.get("invested_loan"), integer=True),
                    "invested_deposit_won": clean_number(aum.get("invested_deposit"), integer=True),
                    "invested_aum": clean_number(aum.get("invested_aum"), integer=True),
                    "aum_status": clean_text(aum.get("status")),
                    "aum_source": paths["aum"].name,
                }
            )
        records.append(record)

        if aum:
            snapshot_date = clean_date(aum.get("input_date")) or clean_date(aum.get("base_date"))
            snapshot_amounts = {
                "aum": clean_number(aum.get("aum"), integer=True),
                "loan": clean_number(aum.get("loan"), integer=True),
                "equity": clean_number(aum.get("equity"), integer=True),
                "deposit": clean_number(aum.get("deposit"), integer=True),
            }
            if snapshot_date and any(value is not None for value in snapshot_amounts.values()):
                snapshots.append(
                    {
                        "snapshot_id": make_id("aum_current", fund_id, snapshot_date),
                        "fund_id": fund_id,
                        "snapshot_date": snapshot_date,
                        "snapshot_year": int(snapshot_date[:4]),
                        "region": None,
                        "sector": record["sector"],
                        **snapshot_amounts,
                        "is_liquidated": record["status"] == "청산",
                        "source_system": "current_aum_snapshot",
                        "metadata": {
                            "source_file": paths["aum"].name,
                            "source_date": paths["aum"].stem.rsplit("_", 1)[-1],
                            "fund_name": fund_name,
                            "short_name": record["short_name"],
                            "asset_name": record["asset_name"],
                            "status": record["status"],
                        },
                    }
                )
            lifecycle.append(
                {
                    "fund_id": fund_id,
                    "op_status": record["status"],
                    "setup_date": record["setup_date"],
                    "maturity_date": record["maturity_date"],
                    "liquidation_date": record["termination_date"],
                    "fund_name": fund_name,
                    "short_name": record["short_name"],
                    "sector": record["sector"],
                    "asset_name": record["asset_name"],
                    "is_aum_target": yn_bool(row.get("AUM합산대상여부")),
                    "aum_base": clean_number(aum.get("aum"), integer=True),
                    "aum_base_date": clean_date(aum.get("input_date")) or clean_date(aum.get("base_date")),
                    "source_system": "current_aum_snapshot",
                    "metadata": {
                        "source_file": paths["aum"].name,
                        "source_date": paths["aum"].stem.rsplit("_", 1)[-1],
                    },
                }
            )
    return records, snapshots, lifecycle, aum_rows


def find_fund_identifier_conflicts(source_records, existing_rows):
    existing_ids = {clean_text(row.get("fund_id")) for row in existing_rows}
    by_name = defaultdict(list)
    for row in existing_rows:
        fund_name = clean_text(row.get("fund_name"))
        if fund_name:
            by_name[fund_name].append(clean_text(row.get("fund_id")))
    conflicts = []
    for record in source_records:
        if record["fund_id"] in existing_ids:
            continue
        matches = [fund_id for fund_id in by_name.get(record.get("fund_name"), []) if fund_id != record["fund_id"]]
        if matches:
            conflicts.append(
                {
                    "fund_id": record["fund_id"],
                    "fund_name": record.get("fund_name"),
                    "reason": "new fund code has an exact existing fund name",
                    "existing_fund_ids": matches,
                }
            )
    return conflicts


def build_asset_master_records(source_rows, existing_rows, source_path):
    code_groups = defaultdict(list)
    for existing in existing_rows:
        existing_code = clean_text(existing.get("asset_code"))
        if existing_code:
            code_groups[existing_code].append(existing)
    by_name_address, _ = unique_index(
        existing_rows,
        lambda row: (clean_text(row.get("canonical_name")), clean_text(row.get("address_text"))),
    )
    by_name, _ = unique_index(existing_rows, lambda row: clean_text(row.get("canonical_name")))
    by_address, _ = unique_index(existing_rows, lambda row: clean_text(row.get("address_text")))
    records = []
    resolution = {}
    conflicts = []
    seen_codes = set()
    for row in source_rows:
        code = clean_text(row.get("자산코드"))
        name = clean_text(row.get("자산(건물)명"))
        address = clean_text(row.get("전체주소(시/도, 구/군 포함)"))
        if not code or not name:
            continue
        if code in seen_codes:
            conflicts.append({"asset_code": code, "reason": "duplicate source asset code"})
            continue
        seen_codes.add(code)
        code_matches = code_groups.get(code, [])
        existing = None
        match_basis = None
        if len(code_matches) == 1:
            existing = code_matches[0]
            match_basis = "asset_code"
        elif len(code_matches) > 1:
            deterministic_id = make_asset_id(code)
            deterministic = [item for item in code_matches if item.get("asset_id") == deterministic_id]
            if len(deterministic) == 1:
                existing = deterministic[0]
                match_basis = "asset_code+deterministic_asset_id"
            else:
                exact_rows = [
                    item
                    for item in code_matches
                    if clean_text(item.get("canonical_name")) == name
                    and clean_text(item.get("address_text")) == address
                ]
                if len(exact_rows) == 1:
                    existing = exact_rows[0]
                    match_basis = "asset_code+asset_name+address"
                else:
                    conflicts.append(
                        {
                            "asset_code": code,
                            "asset_name": name,
                            "reason": "duplicate server asset code without one deterministic exact row",
                            "asset_ids": sorted(item["asset_id"] for item in code_matches),
                        }
                    )
                    continue
        else:
            candidates = []
            if name and address and (name, address) in by_name_address:
                candidates.append((by_name_address[(name, address)], "asset_name+address"))
            if name in by_name:
                candidates.append((by_name[name], "asset_name"))
            if address and address in by_address:
                candidates.append((by_address[address], "address"))
            candidate_ids = {candidate[0]["asset_id"] for candidate in candidates}
            if len(candidate_ids) > 1:
                conflicts.append(
                    {
                        "asset_code": code,
                        "asset_name": name,
                        "reason": "exact identifiers point to different server assets",
                        "asset_ids": sorted(candidate_ids),
                    }
                )
                continue
            existing = candidates[0][0] if candidates else None
            match_basis = "+".join(sorted({basis for _, basis in candidates})) if candidates else "new_asset_code"
        asset_id = existing["asset_id"] if existing else make_asset_id(code)
        record = {
            "asset_id": asset_id,
            "asset_code": code,
            "canonical_name": name,
            "address_text": address,
            "asset_type": clean_text(row.get("기초자산")),
            "business_stage": clean_text(row.get("사업단계")),
            "site_area": clean_number(row.get("토지면적(㎡)")),
            "gross_floor_area": clean_number(row.get("연면적(m²)")),
            "floors_down": clean_number(row.get("건물규모(지하 층수)"), integer=True),
            "floors_up": clean_number(row.get("건물규모(지상 층수)"), integer=True),
            "completion_date": clean_date(row.get("준공(예정)일")),
            "parking": clean_text(row.get("주차대수")),
            "portfolio_region": clean_text(row.get("투자지역")),
            "city": clean_text(row.get("투자도시")),
        }
        if not existing:
            record["review_status"] = "verified"
            record["metadata"] = {
                "source": source_path.name,
                "source_date": source_path.stem.rsplit("_", 1)[-1],
            }
        records.append(record)
        resolution[code] = {"asset_id": asset_id, "name": name, "address": address, "basis": match_basis}
    return records, resolution, conflicts


def build_asset_resolver(asset_resolution, existing_assets):
    source_rows = [
        {
            "asset_id": item["asset_id"],
            "canonical_name": item["name"],
            "address_text": item["address"],
            "asset_code": code,
        }
        for code, item in asset_resolution.items()
    ]
    source_by_name_address, _ = unique_index(
        source_rows,
        lambda row: (clean_text(row.get("canonical_name")), clean_text(row.get("address_text"))),
    )
    source_by_name, _ = unique_index(source_rows, lambda row: clean_text(row.get("canonical_name")))
    source_by_address, _ = unique_index(source_rows, lambda row: clean_text(row.get("address_text")))

    source_codes = set(asset_resolution)
    fallback_rows = [
        row for row in existing_assets if clean_text(row.get("asset_code")) not in source_codes
    ]
    fallback_by_name_address, _ = unique_index(
        fallback_rows,
        lambda row: (clean_text(row.get("canonical_name")), clean_text(row.get("address_text"))),
    )
    fallback_by_name, _ = unique_index(fallback_rows, lambda row: clean_text(row.get("canonical_name")))
    fallback_by_address, _ = unique_index(fallback_rows, lambda row: clean_text(row.get("address_text")))

    def resolve(name, address=None):
        name = clean_text(name)
        address = clean_text(address)
        if name and address and (name, address) in source_by_name_address:
            return source_by_name_address[(name, address)]["asset_id"], "exact_source_asset_name+address"
        if name and name in source_by_name:
            return source_by_name[name]["asset_id"], "exact_source_asset_name"
        if address and address in source_by_address:
            return source_by_address[address]["asset_id"], "exact_source_address"
        if name and address and (name, address) in fallback_by_name_address:
            return fallback_by_name_address[(name, address)]["asset_id"], "exact_asset_name+address"
        if name and name in fallback_by_name:
            return fallback_by_name[name]["asset_id"], "exact_asset_name"
        if address and address in fallback_by_address:
            return fallback_by_address[address]["asset_id"], "exact_address"
        return None, None

    return resolve


def build_fund_asset_records(source_rows, existing_rows, resolve_asset, source_path):
    grouped = defaultdict(list)
    source_address_counts = Counter()
    for row in source_rows:
        fund_id = clean_text(row.get("펀드코드"))
        asset_name = clean_text(row.get("자산(건물)명"))
        if valid_fund_id(fund_id) and asset_name:
            grouped[(fund_id, asset_name)].append(row)
            address = clean_text(row.get("전체주소(시/도, 구/군 포함)"))
            if address:
                source_address_counts[(fund_id, address)] += 1

    existing_by_key, db_collisions = unique_index(
        existing_rows,
        lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("asset_name"))),
    )
    existing_by_fund_address, _ = unique_index(
        existing_rows,
        lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("address"))),
    )
    records = []
    source_conflicts = []
    collapsed_duplicates = []
    resolution = {}
    for key, rows in grouped.items():
        fund_id, asset_name = key
        if str(key) in db_collisions:
            source_conflicts.append({"key": key, "reason": "duplicate server fund asset key"})
            continue
        addresses = {clean_text(row.get("전체주소(시/도, 구/군 포함)")) for row in rows}
        addresses.discard(None)
        semantic_fields = [
            "기초자산",
            "사업단계",
            "자산성격",
            "보유형태",
            "투자전략",
            "Vehicle구분",
            "국내/해외(자산)",
            "투자지역",
            "전체주소(시/도, 구/군 포함)",
            "연면적(m²)",
            "준공(예정)일",
        ]
        signatures = {
            tuple(clean_text(row.get(field)) for field in semantic_fields)
            for row in rows
        }
        if len(signatures) > 1 or len(addresses) > 1:
            source_conflicts.append({"key": key, "reason": "conflicting duplicate source rows"})
            continue
        if len(rows) > 1:
            collapsed_duplicates.append({"key": key, "rows": len(rows)})
        row = rows[0]
        address = clean_text(row.get("전체주소(시/도, 구/군 포함)"))
        existing = existing_by_key.get(key)
        if not existing and address and source_address_counts[(fund_id, address)] == 1:
            existing = existing_by_fund_address.get((fund_id, address))
        asset_id, match_basis = resolve_asset(asset_name, address)
        metadata = merge_metadata(
            existing,
            {
                "source_file": source_path.name,
                "source_date": source_path.stem.rsplit("_", 1)[-1],
                "fund_short_name": clean_text(row.get("약칭")),
                "fund_name": clean_text(row.get("펀드명")),
                "fund_location": clean_text(row.get("국내/해외")),
                "fund_type": clean_text(row.get("펀드유형")),
                "asset_sequence": distinct_join(item.get("순번") for item in rows),
                "base_asset_class": clean_text(row.get("기초자산")),
                "business_stage_class": clean_text(row.get("사업단계")),
                "asset_nature_class": clean_text(row.get("자산성격")),
                "holding_type_class": clean_text(row.get("보유형태")),
                "investment_strategy_class": clean_text(row.get("투자전략")),
                "vehicle_class": clean_text(row.get("Vehicle구분")),
                "asset_location_type": clean_text(row.get("국내/해외(자산)")),
                "investment_country": clean_text(row.get("투자국가")),
                "local_asset_manager": clean_text(row.get("현지운용사")),
                "local_pm": clean_text(row.get("현지PM")),
                "fund_status": clean_text(row.get("운용상태")),
                "fund_shape": clean_text(row.get("펀드형태")),
                "parent_child_type": clean_text(row.get("모자구분")),
                "division": clean_text(row.get("담당부문")),
                "department": clean_text(row.get("담당부서")),
                "manager": clean_text(row.get("담당자")),
                "asset_match_basis": match_basis,
            },
        )
        record = {
            "fund_id": fund_id,
            "asset_name": asset_name,
            "asset_type": clean_text(row.get("기초자산")),
            "address": address,
            "location_category": clean_text(row.get("투자지역")) or clean_text(row.get("국내/해외(자산)")),
            "completion_date": clean_date(row.get("준공(예정)일")),
            "gross_floor_area": clean_postgres_real(row.get("연면적(m²)")),
            "gfa": clean_number(row.get("연면적(m²)")),
            "main_usage": clean_text(row.get("기초자산")),
            "metadata": {key: value for key, value in metadata.items() if value is not None},
        }
        if existing:
            record["id"] = existing["id"]
        if asset_id:
            record["asset_id"] = asset_id
        records.append(record)
        resolution[(fund_id, asset_name)] = {
            "asset_id": asset_id,
            "basis": match_basis,
            "address": address,
        }
    return records, resolution, source_conflicts, collapsed_duplicates, db_collisions


def collapse_exposures(rows, party_column, mapping):
    grouped = defaultdict(list)
    for row in rows:
        fund_id = clean_text(row.get("펀드코드"))
        party = clean_text(row.get(party_column))
        if valid_fund_id(fund_id) and party:
            grouped[(fund_id, party)].append(row)
    records = []
    conflicts = []
    for key, values in grouped.items():
        base_dates = {clean_date(row.get("기준일자")) for row in values}
        base_dates.discard(None)
        if len(base_dates) > 1:
            conflicts.append({"key": key, "reason": "multiple base dates in one source snapshot"})
            continue
        record = {"fund_id": key[0], mapping["raw"]: key[1], mapping["clean"]: key[1]}
        for target, source in mapping.get("sum", {}).items():
            record[target] = sum_or_none(row.get(source) for row in values)
        for target, source in mapping.get("number", {}).items():
            record[target] = first_non_empty((row.get(source) for row in values), clean_number)
        for target, source in mapping.get("date", {}).items():
            record[target] = first_non_empty((row.get(source) for row in values), clean_date)
        for target, source in mapping.get("text", {}).items():
            record[target] = distinct_join(row.get(source) for row in values)
        record["base_date"] = next(iter(base_dates), None)
        record["source_asset_names"] = [clean_text(row.get("자산")) for row in values]
        records.append(record)
    return records, conflicts


def resolve_exposure_asset(record, fund_asset_resolution):
    candidates = set()
    for asset_name in record.pop("source_asset_names", []):
        key = (record["fund_id"], clean_text(asset_name))
        resolved = fund_asset_resolution.get(key)
        if resolved and resolved.get("asset_id"):
            candidates.add(resolved["asset_id"])
    if len(candidates) == 1:
        return next(iter(candidates))
    return None


def build_exposure_records(sources, db, fund_asset_resolution):
    lender_mapping = {
        "raw": "lender_raw",
        "clean": "lender_clean",
        "sum": {
            "committed_amt": "대출약정금액(원)",
            "drawn_amt": "대출인출금액(원)",
            "remaining_amt": "대출잔여금액(원)",
        },
        "number": {
            "base_rate": "기준금리",
            "spread_rate": "가산금리",
            "interest_rate": "대출금리",
            "all_in_rate": "All-in금리",
        },
        "date": {"drawdown_date": "대출인출일", "loan_maturity_date": "대출만기일"},
        "text": {"trench": "트렌치", "interest_type": "이자유형", "remarks": "비고"},
    }
    beneficiary_mapping = {
        "raw": "beneficiary_raw",
        "clean": "beneficiary_clean",
        "sum": {
            "committed_amt": "총약정금액",
            "invested_amt": "투입금액",
            "remaining_amt": "잔여약정금액",
            "setup_units": "설정해지좌수",
            "setup_amt": "설정해지금액",
        },
        "number": {"share_ratio": "비율(%)"},
        "date": {"invested_date": "약정콜일자"},
        "text": {"remarks": "비고"},
    }
    lenders, lender_source_conflicts = collapse_exposures(sources["lender"], "대주", lender_mapping)
    beneficiaries, beneficiary_source_conflicts = collapse_exposures(
        sources["beneficiary"], "수익자", beneficiary_mapping
    )
    lender_index, lender_db_collisions = unique_index(
        db["lender_exposures"],
        lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("lender_clean"))),
    )
    beneficiary_index, beneficiary_db_collisions = unique_index(
        db["beneficiary_exposures"],
        lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("beneficiary_clean"))),
    )

    output_lenders = []
    for record in lenders:
        key = (record["fund_id"], record["lender_clean"])
        if str(key) in lender_db_collisions:
            continue
        existing = lender_index.get(key)
        asset_id = resolve_exposure_asset(record, fund_asset_resolution)
        if existing:
            record["id"] = existing["id"]
        if asset_id:
            record["asset_id"] = asset_id
        output_lenders.append(record)

    output_beneficiaries = []
    for record in beneficiaries:
        key = (record["fund_id"], record["beneficiary_clean"])
        if str(key) in beneficiary_db_collisions:
            continue
        existing = beneficiary_index.get(key)
        asset_id = resolve_exposure_asset(record, fund_asset_resolution)
        if existing:
            record["id"] = existing["id"]
        if asset_id:
            record["asset_id"] = asset_id
        output_beneficiaries.append(record)

    conflicts = {
        "lender_source": lender_source_conflicts,
        "lender_db": lender_db_collisions,
        "beneficiary_source": beneficiary_source_conflicts,
        "beneficiary_db": beneficiary_db_collisions,
    }
    return output_lenders, output_beneficiaries, conflicts


def reconcile_table(existing_rows, expected_rows, existing_key, expected_key):
    existing_index, collisions = unique_index(existing_rows, existing_key)
    inserts = []
    updates = []
    unchanged = []
    skipped = []
    matched_keys = set()
    for expected in expected_rows:
        key = expected_key(expected)
        if str(key) in collisions:
            skipped.append({"key": key, "reason": "server key collision"})
            continue
        existing = existing_index.get(key)
        if not existing:
            inserts.append(expected)
            continue
        matched_keys.add(key)
        differences = changed_fields(existing, expected)
        if differences:
            updates.append({"record": expected, "changed_fields": differences})
        else:
            unchanged.append(expected)
    return {
        "insert_records": inserts,
        "update_records": [item["record"] for item in updates],
        "update_details": [{"key": expected_key(item["record"]), "changed_fields": item["changed_fields"]} for item in updates],
        "unchanged_records": unchanged,
        "skipped": skipped,
        "retained_not_in_source": max(0, len(existing_rows) - len(matched_keys)),
        "server_collisions": collisions,
    }


def plan_summary(result):
    return {
        "insert": len(result["insert_records"]),
        "update": len(result["update_records"]),
        "unchanged": len(result["unchanged_records"]),
        "skipped": len(result["skipped"]),
        "retained_not_in_source": result["retained_not_in_source"],
        "server_collision_groups": len(result["server_collisions"]),
        "changed_field_counts": dict(
            Counter(field for item in result["update_details"] for field in item["changed_fields"])
        ),
    }


def backup_database(db, source_date):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = OUTPUT_DIR / f"fund_db_update_{source_date}_{timestamp}"
    before_dir = backup_dir / "before"
    before_dir.mkdir(parents=True, exist_ok=True)
    for table, rows in db.items():
        (before_dir / f"{table}.json").write_text(
            json.dumps(rows, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
    return backup_dir


def build_link_records(fund_asset_records, current_fund_assets, current_links, source_path):
    fund_asset_index, collisions = unique_index(
        current_fund_assets,
        lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("asset_name"))),
    )
    link_index, link_collisions = unique_index(
        current_links,
        lambda row: (row.get("asset_id"), row.get("fund_id"), row.get("relation_type")),
    )
    links = []
    unresolved = []
    for record in fund_asset_records:
        key = (record["fund_id"], record["asset_name"])
        source_row = fund_asset_index.get(key)
        asset_id = record.get("asset_id")
        if not source_row or not asset_id:
            unresolved.append({"key": key, "reason": "missing exact asset or fund_asset row"})
            continue
        link_key = (asset_id, record["fund_id"], "underlying_asset")
        if str(link_key) in link_collisions:
            unresolved.append({"key": link_key, "reason": "duplicate server link key"})
            continue
        existing = link_index.get(link_key)
        metadata = merge_metadata(
            existing,
            {
                "match_reason": "exact_excel_identifier",
                "source_file": source_path.name,
                "source_date": source_path.stem.rsplit("_", 1)[-1],
            },
        )
        links.append(
            {
                "asset_id": asset_id,
                "fund_id": record["fund_id"],
                "relation_type": "underlying_asset",
                "source_table": "fund_assets",
                "source_id": str(source_row["id"]),
                "confidence": 1.0,
                "metadata": metadata,
                "exposure_role": "direct_owner",
                "directness": "direct",
            }
        )
    return links, unresolved, collisions


def verify_records(actual_rows, expected_rows, actual_key, expected_key):
    actual_index, collisions = unique_index(actual_rows, actual_key)
    mismatches = []
    for expected in expected_rows:
        key = expected_key(expected)
        actual = actual_index.get(key)
        if not actual:
            mismatches.append({"key": key, "reason": "missing"})
            continue
        differences = changed_fields(actual, expected)
        if differences:
            mismatches.append({"key": key, "reason": "value_mismatch", "fields": differences})
    return {"expected": len(expected_rows), "mismatches": mismatches, "collision_groups": collisions}


def main():
    parser = argparse.ArgumentParser(description="Incrementally sync current fund Excel sources to Supabase.")
    parser.add_argument("--source-date", default=datetime.now().strftime("%Y%m%d"))
    parser.add_argument("--apply", action="store_true", help="Apply the prepared incremental update.")
    args = parser.parse_args()

    paths = source_files(args.source_date)
    sources = load_sources(paths)
    url, api_key = read_env()
    client = PostgrestClient(url, api_key)
    print("Fetching live database snapshots...")
    db = {table: client.fetch_all(table) for table in TABLES}
    backup_dir = backup_database(db, args.source_date)

    fund_records, snapshot_records, lifecycle_records, aum_rows = build_fund_records(sources, paths)
    fund_identifier_conflicts = find_fund_identifier_conflicts(fund_records, db["funds"])
    lifecycle_by_id = {row["fund_id"]: row for row in db["fund_lifecycle"]}
    lifecycle_records = [
        {
            **record,
            "metadata": merge_metadata(lifecycle_by_id.get(record["fund_id"]), record["metadata"]),
        }
        for record in lifecycle_records
    ]
    asset_records, asset_resolution, asset_conflicts = build_asset_master_records(
        sources["asset_manage"], db["asset_master"], paths["asset_manage"]
    )
    resolve_asset = build_asset_resolver(asset_resolution, db["asset_master"])
    fund_asset_records, fund_asset_resolution, fund_asset_conflicts, collapsed_duplicates, fund_asset_db_collisions = (
        build_fund_asset_records(
            sources["fund_asset"], db["fund_assets"], resolve_asset, paths["fund_asset"]
        )
    )
    lender_records, beneficiary_records, exposure_conflicts = build_exposure_records(
        sources, db, fund_asset_resolution
    )

    plans = {
        "funds": reconcile_table(db["funds"], fund_records, lambda row: row.get("fund_id"), lambda row: row["fund_id"]),
        "asset_master": reconcile_table(
            db["asset_master"], asset_records, lambda row: row.get("asset_id"), lambda row: row["asset_id"]
        ),
        "fund_assets": reconcile_table(
            db["fund_assets"],
            fund_asset_records,
            lambda row: ("id", row.get("id")),
            lambda row: ("id", row["id"]) if row.get("id") is not None else ("new", row["fund_id"], row["asset_name"]),
        ),
        "lender_exposures": reconcile_table(
            db["lender_exposures"],
            lender_records,
            lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("lender_clean"))),
            lambda row: (row["fund_id"], row["lender_clean"]),
        ),
        "beneficiary_exposures": reconcile_table(
            db["beneficiary_exposures"],
            beneficiary_records,
            lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("beneficiary_clean"))),
            lambda row: (row["fund_id"], row["beneficiary_clean"]),
        ),
        "aum_snapshots": reconcile_table(
            db["aum_snapshots"], snapshot_records, lambda row: row.get("snapshot_id"), lambda row: row["snapshot_id"]
        ),
        "fund_lifecycle": reconcile_table(
            db["fund_lifecycle"], lifecycle_records, lambda row: row.get("fund_id"), lambda row: row["fund_id"]
        ),
    }

    report = {
        "source_date": args.source_date,
        "source_files": {key: path.name for key, path in paths.items()},
        "source_rows": {
            "fund": len(fund_records),
            "aum_raw": len(sources["aum"]),
            "aum_valid": len(aum_rows),
            "asset_manage": len(asset_records),
            "fund_asset": len(fund_asset_records),
            "lender_raw": len(sources["lender"]),
            "lender_collapsed": len(lender_records),
            "beneficiary_raw": len(sources["beneficiary"]),
            "beneficiary_collapsed": len(beneficiary_records),
        },
        "before_counts": {table: len(rows) for table, rows in db.items()},
        "plan": {table: plan_summary(plan) for table, plan in plans.items()},
        "conflicts": {
            "funds": fund_identifier_conflicts,
            "asset_master": asset_conflicts,
            "fund_assets": fund_asset_conflicts,
            "fund_asset_server": fund_asset_db_collisions,
            "fund_asset_collapsed_duplicates": collapsed_duplicates,
            "exposures": exposure_conflicts,
        },
        "matching_policy": {
            "funds": "exact fund_id; exact-name/code conflicts are not fuzzy-merged",
            "asset_master": "exact asset_code, then exact name+address/name/address only when unique",
            "fund_assets": "exact fund_id+asset_name, then exact fund_id+address only when unique",
            "exposures": "exact fund_id+party name; base_date is updated as snapshot data",
            "normalization": "outer whitespace and Excel scalar/date conversion only; no fuzzy matching",
            "deletes": 0,
        },
        "applied": False,
    }
    (backup_dir / "plan.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )

    print(json.dumps({"backup_dir": str(backup_dir), **report}, ensure_ascii=False, indent=2, default=str))
    blocking_conflicts = len(fund_identifier_conflicts) + len(asset_conflicts) + len(fund_asset_conflicts)
    blocking_conflicts += len(exposure_conflicts["lender_source"]) + len(exposure_conflicts["beneficiary_source"])
    if blocking_conflicts:
        print(f"Blocking source/identifier conflicts: {blocking_conflicts}", file=sys.stderr)
        raise SystemExit(2)
    if not args.apply:
        print("Dry-run complete. Re-run with --apply after reviewing plan.json.")
        return

    print("Applying incremental updates (no deletes)...")
    fund_changes = plans["funds"]["insert_records"] + plans["funds"]["update_records"]
    asset_changes = plans["asset_master"]["insert_records"] + plans["asset_master"]["update_records"]
    fund_asset_changes = plans["fund_assets"]["insert_records"] + plans["fund_assets"]["update_records"]
    lender_changes = plans["lender_exposures"]["insert_records"] + plans["lender_exposures"]["update_records"]
    beneficiary_changes = (
        plans["beneficiary_exposures"]["insert_records"]
        + plans["beneficiary_exposures"]["update_records"]
    )
    snapshot_changes = plans["aum_snapshots"]["insert_records"] + plans["aum_snapshots"]["update_records"]
    lifecycle_changes = plans["fund_lifecycle"]["insert_records"] + plans["fund_lifecycle"]["update_records"]

    client.write_grouped("funds", fund_changes, on_conflict="fund_id")
    client.write_grouped("asset_master", asset_changes, on_conflict="asset_id")

    existing_fund_assets = [record for record in fund_asset_changes if record.get("id") is not None]
    new_fund_assets = [{key: value for key, value in record.items() if key != "id"} for record in fund_asset_changes if record.get("id") is None]
    client.write_grouped("fund_assets", existing_fund_assets, on_conflict="id")
    client.write_grouped("fund_assets", new_fund_assets)

    existing_lenders = [record for record in lender_changes if record.get("id") is not None]
    new_lenders = [{key: value for key, value in record.items() if key != "id"} for record in lender_changes if record.get("id") is None]
    client.write_grouped("lender_exposures", existing_lenders, on_conflict="id")
    client.write_grouped("lender_exposures", new_lenders)

    existing_beneficiaries = [record for record in beneficiary_changes if record.get("id") is not None]
    new_beneficiaries = [
        {key: value for key, value in record.items() if key != "id"}
        for record in beneficiary_changes
        if record.get("id") is None
    ]
    client.write_grouped("beneficiary_exposures", existing_beneficiaries, on_conflict="id")
    client.write_grouped("beneficiary_exposures", new_beneficiaries)
    client.write_grouped("aum_snapshots", snapshot_changes, on_conflict="snapshot_id")

    lifecycle_to_write = lifecycle_records
    client.write_grouped("fund_lifecycle", lifecycle_changes, on_conflict="fund_id")

    current_fund_assets = client.fetch_all("fund_assets")
    current_links = client.fetch_all("asset_fund_links")
    link_records, unresolved_links, link_source_collisions = build_link_records(
        fund_asset_records, current_fund_assets, current_links, paths["fund_asset"]
    )
    link_plan = reconcile_table(
        current_links,
        link_records,
        lambda row: (row.get("asset_id"), row.get("fund_id"), row.get("relation_type")),
        lambda row: (row["asset_id"], row["fund_id"], row["relation_type"]),
    )
    link_changes = link_plan["insert_records"] + link_plan["update_records"]
    client.write_grouped(
        "asset_fund_links", link_changes, on_conflict="asset_id,fund_id,relation_type"
    )
    client.request("POST", "rpc/refresh_party_exposure_surfaces", payload={})

    after = {table: client.fetch_all(table) for table in TABLES}
    verification = {
        "funds": verify_records(after["funds"], fund_records, lambda row: row.get("fund_id"), lambda row: row["fund_id"]),
        "asset_master": verify_records(
            after["asset_master"], asset_records, lambda row: row.get("asset_id"), lambda row: row["asset_id"]
        ),
        "fund_assets": verify_records(
            after["fund_assets"],
            fund_asset_records,
            lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("asset_name"))),
            lambda row: (row["fund_id"], row["asset_name"]),
        ),
        "lender_exposures": verify_records(
            after["lender_exposures"],
            lender_records,
            lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("lender_clean"))),
            lambda row: (row["fund_id"], row["lender_clean"]),
        ),
        "beneficiary_exposures": verify_records(
            after["beneficiary_exposures"],
            beneficiary_records,
            lambda row: (clean_text(row.get("fund_id")), clean_text(row.get("beneficiary_clean"))),
            lambda row: (row["fund_id"], row["beneficiary_clean"]),
        ),
        "aum_snapshots": verify_records(
            after["aum_snapshots"], snapshot_records, lambda row: row.get("snapshot_id"), lambda row: row["snapshot_id"]
        ),
        "fund_lifecycle": verify_records(
            after["fund_lifecycle"], lifecycle_to_write, lambda row: row.get("fund_id"), lambda row: row["fund_id"]
        ),
        "asset_fund_links": verify_records(
            after["asset_fund_links"],
            link_records,
            lambda row: (row.get("asset_id"), row.get("fund_id"), row.get("relation_type")),
            lambda row: (row["asset_id"], row["fund_id"], row["relation_type"]),
        ),
    }
    mismatch_count = sum(len(item["mismatches"]) for item in verification.values())
    final_report = {
        **report,
        "applied": True,
        "after_counts": {table: len(rows) for table, rows in after.items()},
        "links": {
            "prepared": len(link_records),
            "plan": plan_summary(link_plan),
            "unresolved": unresolved_links,
            "source_collision_groups": link_source_collisions,
        },
        "verification": verification,
        "verification_mismatch_count": mismatch_count,
    }
    (backup_dir / "result.json").write_text(
        json.dumps(final_report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    print(json.dumps(final_report, ensure_ascii=False, indent=2, default=str))
    if mismatch_count:
        raise SystemExit(f"Verification failed with {mismatch_count} mismatches")
    print("Incremental update and live read-back verification completed.")


if __name__ == "__main__":
    main()
