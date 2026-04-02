from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parent
WORKBOOK_GLOB = "20260401_*좌석배치*.xlsx"
OUTPUT_JSON = ROOT / "seat-layout-data.json"
OUTPUT_JS = ROOT / "seat-layout-data.js"
OVERLAY_JS = ROOT / "seat-layout-overlays.js"

SEAT_CODE_RE = re.compile(r"^[A-Z]{1,3}\d{1,3}$")

SPACE_KEYWORDS = (
    "회의실",
    "화장실",
    "라운지",
    "엘리베이터",
    "계단실",
    "서버실",
    "창고",
    "OA",
    "그룹장실",
    "파트장실",
    "집무실",
    "사무실",
    "홀",
    "폰룸",
    "락커룸",
    "카페",
    "리셉션",
    "대기실",
    "전략리서치실",
    "운용지원실",
    "안전보건관리실",
)


@dataclass(frozen=True)
class SheetMeta:
    floor_code: str
    scenario: str


def normalize_text(value: Any) -> str:
    text = str(value or "").replace("\n", " ").replace("\r", " ")
    return " ".join(text.split()).strip()


def infer_sheet_meta(sheet_name: str) -> SheetMeta | None:
    floor_match = re.search(r"(2F|12F|13F)", sheet_name)
    if not floor_match:
        return None
    scenario = "plan" if sheet_name.endswith("_1") else "current"
    return SheetMeta(floor_code=floor_match.group(1), scenario=scenario)


def color_to_hex(cell: Cell) -> str | None:
    fill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return None
    rgb = fill.fgColor.rgb
    if not rgb:
        return None
    rgb = str(rgb)
    if not re.fullmatch(r"[0-9A-Fa-f]{8}|[0-9A-Fa-f]{6}", rgb):
        return None
    rgb = rgb[-6:].upper()
    if rgb in {"000000", "FFFFFF"}:
        return None
    return f"#{rgb}"


def classify_shape(label: str) -> str:
    if "회의실" in label:
        return "room"
    if any(keyword in label for keyword in ("화장실", "엘리베이터", "계단실", "홀", "락커룸", "카페", "리셉션", "대기실")):
        return "common"
    if any(keyword in label for keyword in ("라운지", "OA", "창고", "서버실")):
        return "support"
    if any(keyword in label for keyword in ("그룹장실", "파트장실", "집무실")):
        return "office"
    if label:
        return "zone"
    return "block"


def is_seat_code(text: str) -> bool:
    return bool(SEAT_CODE_RE.match(text))


def build_merged_index(ws: Worksheet) -> dict[tuple[int, int], str]:
    merged_index: dict[tuple[int, int], str] = {}
    for merged in ws.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_index[(row, col)] = str(merged)
    return merged_index


def extract_shapes(ws: Worksheet, floor_code: str) -> list[dict[str, Any]]:
    shapes: list[dict[str, Any]] = []
    seen_ranges: set[str] = set()

    for merged in ws.merged_cells.ranges:
        top_left = ws.cell(merged.min_row, merged.min_col)
        label = normalize_text(top_left.value)
        fill = color_to_hex(top_left)
        if not label and not fill:
            continue
        seen_ranges.add(str(merged))
        shape_id = f"{floor_code}-shape-{len(shapes) + 1:03d}"
        shapes.append(
            {
                "shapeId": shape_id,
                "floorCode": floor_code,
                "shapeType": classify_shape(label),
                "label": label,
                "x": merged.min_col - 1,
                "y": merged.min_row - 1,
                "w": merged.max_col - merged.min_col + 1,
                "h": merged.max_row - merged.min_row + 1,
                "fill": fill,
                "source": str(merged),
            }
        )

    return shapes


def maybe_name_below(ws: Worksheet, row: int, col: int) -> str:
    for next_row in (row + 1, row + 2):
        text = normalize_text(ws.cell(next_row, col).value)
        if not text:
            continue
        if is_seat_code(text):
            return ""
        if any(keyword in text for keyword in SPACE_KEYWORDS):
            return ""
        return text
    return ""


def extract_floor_data(ws: Worksheet, meta: SheetMeta) -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, Any]]:
    merged_index = build_merged_index(ws)
    seats: list[dict[str, Any]] = []
    assignments: dict[str, dict[str, str]] = {}

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            text = normalize_text(cell.value)
            if not text or not is_seat_code(text):
                continue
            if (row, col) in merged_index:
                continue

            seat_code = f"{meta.floor_code}-{text}"
            seat = {
                "seatCode": seat_code,
                "seatLabel": text,
                "floorCode": meta.floor_code,
                "x": col - 1,
                "y": row - 1,
                "w": 1,
                "h": 1,
                "sourceCell": cell.coordinate,
            }
            seats.append(seat)
            assignments[seat_code] = {
                "personName": maybe_name_below(ws, row, col),
                "sheetName": ws.title,
            }

    bounds = {
        "floorCode": meta.floor_code,
        "sheetName": ws.title,
        "scenario": meta.scenario,
        "rows": ws.max_row,
        "cols": ws.max_column,
    }
    return seats, assignments, bounds


def merge_seat_defs(existing: dict[str, dict[str, Any]], incoming: list[dict[str, Any]]) -> None:
    for seat in incoming:
        current = existing.get(seat["seatCode"])
        if not current:
            existing[seat["seatCode"]] = seat
            continue
        # Keep earliest coordinates unless later sheet gives a more complete one.
        if current["x"] != seat["x"] or current["y"] != seat["y"]:
            current["x"] = min(current["x"], seat["x"])
            current["y"] = min(current["y"], seat["y"])


def create_overlay_stub(floors: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "zones": [],
        "externalSeatCodes": {
            floor["floorCode"]: [] for floor in floors
        },
        "notes": "externalSeatCodes에 좌석코드를 넣으면 회색 처리할 수 있습니다. zones는 seatCode 목록 기준으로 구역 오버레이를 정의합니다.",
    }


def build_payload() -> dict[str, Any]:
    workbook_path = next(ROOT.glob(WORKBOOK_GLOB))
    wb = load_workbook(workbook_path, data_only=True)

    floor_index: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        meta = infer_sheet_meta(ws.title)
        if not meta:
            continue

        floor = floor_index.setdefault(
            meta.floor_code,
            {
                "floorCode": meta.floor_code,
                "displayName": meta.floor_code,
                "scenarios": {},
                "shapes": [],
                "seatDefs": {},
            },
        )

        shapes = extract_shapes(ws, meta.floor_code)
        seats, assignments, bounds = extract_floor_data(ws, meta)

        if not floor["shapes"]:
            floor["shapes"] = shapes
        merge_seat_defs(floor["seatDefs"], seats)

        floor["scenarios"][meta.scenario] = {
            "assignments": assignments,
            "bounds": bounds,
        }

    floors = []
    for floor_code in sorted(floor_index.keys(), key=lambda value: int(value.replace("F", ""))):
        floor = floor_index[floor_code]
        floors.append(
            {
                "floorCode": floor["floorCode"],
                "displayName": floor["displayName"],
                "shapes": floor["shapes"],
                "seatDefs": sorted(floor["seatDefs"].values(), key=lambda seat: (seat["y"], seat["x"])),
                "scenarios": floor["scenarios"],
            }
        )

    payload = {
        "meta": {
            "sourceWorkbook": workbook_path.name,
            "generatedAt": __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "floorOrder": [floor["floorCode"] for floor in floors],
        },
        "floors": floors,
    }
    return payload


def write_outputs(payload: dict[str, Any]) -> None:
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.SEAT_LAYOUT_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    if not OVERLAY_JS.exists():
        OVERLAY_JS.write_text(
            "window.SEAT_LAYOUT_OVERLAYS = "
            + json.dumps(create_overlay_stub(payload["floors"]), ensure_ascii=False, indent=2)
            + ";\n",
            encoding="utf-8",
        )


def main() -> None:
    payload = build_payload()
    write_outputs(payload)
    print(f"Generated {OUTPUT_JSON.name} and {OUTPUT_JS.name}")


if __name__ == "__main__":
    main()
