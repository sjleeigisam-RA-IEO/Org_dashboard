import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parents[1]
JSON_PATH = BASE / "seat-layout-data.json"
OUT_PATH = BASE / "seat_layout_google_sheet_upload.xlsx"

MOVE_OVERRIDES = {
    "2F-B19": {"ignore": True},
    "2F-B65": {"force_origin_floor": "12F", "force_origin_seat": ""},
}

OTHER_DEPT_PREFIX = {
    "2F": "2F-A",
    "12F": "12F-B",
}

FLOOR_ORDER = ["2F", "12F", "13F"]


def natural_seat_key(seat_code: str):
    match = re.match(r"^(\d+F)-([A-Z]+)(\d+)$", seat_code or "")
    if not match:
        return (999, "Z", 9999)
    return (int(match.group(1).replace("F", "")), match.group(2), int(match.group(3)))


def load_data():
    with JSON_PATH.open("r", encoding="utf-8-sig") as file:
        return json.load(file)


def build_rows(data):
    floors = {floor["floorCode"]: floor for floor in data["floors"]}

    current_by_name = {}
    for floor in floors.values():
        assignments = floor.get("scenarios", {}).get("current", {}).get("assignments", {})
        for seat_code, info in assignments.items():
            person = (info or {}).get("personName", "").strip()
            if person:
                current_by_name[person] = seat_code

    rows = []
    for floor_code in FLOOR_ORDER:
        floor = floors.get(floor_code)
        if not floor:
            continue
        plan_assignments = floor.get("scenarios", {}).get("plan", {}).get("assignments", {})
        for seat in sorted(floor.get("seatDefs", []), key=lambda item: natural_seat_key(item.get("seatCode", ""))):
            seat_code = seat["seatCode"]
            info = plan_assignments.get(seat_code, {}) or {}
            person = (info.get("personName") or "").strip()
            override = MOVE_OVERRIDES.get(seat_code, {})

            origin_seat = ""
            origin_floor = ""
            is_moved = "N"
            note = ""

            if override.get("ignore"):
                note = "동명이인 예외: 이동 표시 제외"
            elif person:
                origin_seat = current_by_name.get(person, "")
                if "force_origin_seat" in override:
                    origin_seat = override.get("force_origin_seat", "")
                origin_floor = override.get("force_origin_floor") or (origin_seat.split("-")[0] if origin_seat else "")
                if origin_floor or origin_seat:
                    is_moved = "Y" if origin_seat != seat_code else "N"
                    if origin_seat == "" and origin_floor:
                        is_moved = "Y"
                if seat_code == "2F-B65":
                    note = "동명이인 예외: 원래 12층으로만 표시"

            is_external = "Y" if seat_code.startswith(OTHER_DEPT_PREFIX.get(floor_code, "__NONE__")) else "N"

            rows.append(
                {
                    "seat_code": seat_code,
                    "floor_code": floor_code,
                    "seat_label": seat.get("seatLabel", ""),
                    "person_name": person,
                    "origin_floor_code": origin_floor,
                    "origin_seat_code": origin_seat,
                    "is_moved": is_moved,
                    "is_external_division": is_external,
                    "note": note,
                }
            )
    return rows


def build_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "seat_layout_latest"

    headers = [
        "seat_code",
        "floor_code",
        "seat_label",
        "person_name",
        "origin_floor_code",
        "origin_seat_code",
        "is_moved",
        "is_external_division",
        "note",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col, width in {"A": 16, "B": 10, "C": 10, "D": 14, "E": 12, "F": 16, "G": 10, "H": 18, "I": 32}.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    guide = workbook.create_sheet("guide")
    guide_rows = [
        ["목적", "구글시트 새 탭에 붙여넣어 자리배치 최신 상태를 관리하는 입력 원본"],
        ["기준", "한 행 = 한 좌석"],
        ["필수 수정 컬럼", "person_name, origin_floor_code, origin_seat_code, is_moved, is_external_division, note"],
        ["seat_code", "불변 좌석코드. 수정하지 않음"],
        ["floor_code", "층 코드. 수정하지 않음"],
        ["person_name", "현재 그 자리에 앉는 사람 이름"],
        ["origin_floor_code", "이동자면 원래 층 (예: 12F)"],
        ["origin_seat_code", "같은 층 이동 또는 정확한 원래 좌석을 쓸 때 입력"],
        ["is_moved", "이동이면 Y, 아니면 N"],
        ["is_external_division", "타 부문 좌석이면 Y, 아니면 N"],
        ["note", "동명이인 등 예외 메모"],
    ]
    for row in guide_rows:
        guide.append(row)

    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font

    for col, width in {"A": 28, "B": 72}.items():
        guide.column_dimensions[col].width = width

    workbook.save(OUT_PATH)


def verify():
    workbook = load_workbook(OUT_PATH, data_only=True)
    guide = workbook["guide"]
    sheet = workbook["seat_layout_latest"]
    print(OUT_PATH)
    print("A2:", guide["A2"].value)
    print("B2:", guide["B2"].value)
    print("A3:", guide["A3"].value)
    print("B3:", guide["B3"].value)
    print("sample seat:", sheet["A2"].value, sheet["D2"].value, sheet["I2"].value)
    print("rows:", sheet.max_row - 1)


if __name__ == "__main__":
    payload = load_data()
    rows = build_rows(payload)
    build_workbook(rows)
    verify()
