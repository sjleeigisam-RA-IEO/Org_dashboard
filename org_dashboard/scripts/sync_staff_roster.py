from __future__ import annotations

import argparse
import hashlib
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DASHBOARD_DIR = ROOT / "org_dashboard"
DEFAULT_SOURCE = Path(
    r"C:\10137_WorkSpace\00. 2025 RA 기획추진\03. 부문 내 업무"
    r"\00. 부문데이터\구성원 데이터\★리얼에셋부문 인력현황_260701기준.xlsx"
)
DEFAULT_AS_OF = date(2026, 7, 1)
KST = ZoneInfo("Asia/Seoul")

SECTION_ORDER = ["투자+펀딩", "사업+개발", "관리+운영", "부문직속", "TFs"]
GROUP_ORDER = {
    "투자+펀딩": ["투자1그룹", "투자2그룹", "투자3그룹", "글로벌투자그룹", "스페셜시츄에이션그룹"],
    "사업+개발": ["사업그룹", "디지털사업그룹", "개발솔루션센터"],
    "관리+운영": ["국내자산관리그룹", "글로벌자산관리그룹", "리빙그룹", "리테일솔루션센터"],
    "부문직속": ["리얼에셋부문", "론파이낸스센터", "기업마케팅센터", "공간솔루션센터", "기획추진센터"],
    "TFs": ["SS&C TF", "IOTA CFT", "개발PFV TF"],
}
GROUP_SECTION = {
    group: section
    for section, groups in GROUP_ORDER.items()
    if section != "TFs"
    for group in groups
}
GROUP_SECTION.update({"투자&펀딩": "투자+펀딩", "관리&운영": "관리+운영"})

ROLE_ORDER = ["그룹장", "파트장/센터장", "담당디렉터", "시니어매니저", "매니저"]
LEADER_LABELS = {"부문대표", "부대표", "그룹장", "파트장", "센터장"}
IGNORED_CONCURRENT_TOKENS = {"SMP", "사업&개발", "투자&펀딩", "관리&운영"}


@dataclass(frozen=True)
class OrgPath:
    section: str
    group: str
    part: str
    team: str

    @property
    def key(self) -> str:
        return "|".join((self.section, self.group, self.part, self.team))

    @property
    def display_path(self) -> str:
        return " > ".join(value for value in (self.section, self.group, self.part, self.team) if value)

    def group_scope(self) -> "OrgPath":
        return replace(self, part="", team="")

    def part_scope(self) -> "OrgPath":
        return replace(self, team="")


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return "" if text == "-" else re.sub(r"\s+", " ", text)


def normalize_email(value: object) -> str:
    return clean(value).lower()


def stable_id(prefix: str, *parts: str, length: int = 16) -> str:
    digest = hashlib.sha1("::".join(parts).encode("utf-8")).hexdigest()[:length]
    return f"{prefix}_{digest}"


def to_iso_date(value: object) -> str:
    if value in (None, "", "-"):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean(value)


def normalize_group_name(value: str) -> str:
    value = clean(value)
    return "론파이낸스센터" if value == "Loan Finance센터" else value


def normalize_part_name(raw_part: str, group: str) -> str:
    part = clean(raw_part)
    if not part:
        return "미지정"
    for prefix in (group, "글로벌자산관리그룹", "국내자산관리그룹", "스페셜시츄에이션그룹", "사업그룹"):
        if prefix and part.startswith(prefix):
            part = part[len(prefix) :]
    if group == "국내자산관리그룹" and part.startswith("자산관리"):
        part = part[len("자산관리") :]
    if group == "글로벌자산관리그룹" and part.startswith("글로벌자산관리"):
        part = part[len("글로벌자산관리") :]
    if part == "포트폴리오관리":
        return "포트폴리오 / 관리"
    return part or "미지정"


def normalize_team_name(raw_team: str, raw_org: str, group: str, part: str) -> str:
    team = clean(raw_team) or group
    if team == "로지스틱스매니지먼트":
        return "로지스틱스 / 매니지먼트"
    if team == "리빙매니지먼트":
        return "리빙 / 매니지먼트"
    if team == "포트폴리오관리":
        return "포트폴리오 / 관리"

    compact = team
    for prefix in (group, "글로벌자산관리그룹", "국내자산관리그룹"):
        if prefix and compact.startswith(prefix):
            compact = compact[len(prefix) :]
    if group == "국내자산관리그룹" and compact.startswith("자산관리"):
        compact = compact[len("자산관리") :]
    if group == "글로벌자산관리그룹" and compact.startswith("글로벌자산관리"):
        compact = compact[len("글로벌자산관리") :]

    if part != "미지정" and compact.startswith(part):
        suffix = compact[len(part) :]
        if suffix == "펀드지원":
            return "펀드지원"
        if not suffix:
            return part
        if suffix.isdigit():
            return f"{part}{suffix}"

    if compact in {"", raw_org, group}:
        return part if part != "미지정" else group
    return compact


def split_group_and_part(raw_org: str) -> tuple[str, str]:
    raw_org = clean(raw_org)
    if "/" in raw_org:
        group, raw_part, *_ = [clean(value) for value in raw_org.split("/")]
        return normalize_group_name(group), raw_part

    known_groups = sorted(GROUP_SECTION, key=len, reverse=True)
    for group in known_groups:
        if raw_org.startswith(group) and raw_org != group:
            suffix = raw_org[len(group) :]
            if suffix.endswith("파트"):
                return normalize_group_name(group), suffix
    return normalize_group_name(raw_org), ""


def parse_org_path(raw_org: str, raw_team: str) -> OrgPath:
    raw_org = clean(raw_org)
    raw_team = clean(raw_team)
    if raw_org == "리얼에셋부문":
        return OrgPath("부문직속", "리얼에셋부문", "미지정", raw_team or "리얼에셋부문")
    if raw_org in {"투자&펀딩", "관리&운영"}:
        return OrgPath(GROUP_SECTION[raw_org], raw_org, "미지정", raw_team or raw_org)

    group, raw_part = split_group_and_part(raw_org)
    if group not in GROUP_SECTION:
        raise ValueError(f"지원하지 않는 조직명: {raw_org}")
    section = GROUP_SECTION[group]

    if group.endswith("센터"):
        return OrgPath(section, group, "미지정", normalize_group_name(raw_team or group))

    if not raw_part and group == "스페셜시츄에이션그룹" and raw_team not in {"", group}:
        part = normalize_part_name(raw_team, group)
        return OrgPath(section, group, part, normalize_team_name(raw_team, raw_org, group, part))

    part = normalize_part_name(raw_part, group)
    team = normalize_team_name(raw_team, raw_org, group, part)
    return OrgPath(section, group, part, team)


def classify_role(row: dict[str, Any]) -> tuple[str, str, list[str]]:
    raw_role = clean(row.get("직책(26년)"))
    tags: list[str] = []
    if "대행" in raw_role:
        tags.append("대행")
    if clean(row.get("직위(26년)")) == "사원" or "인턴" in " ".join(clean(value) for value in row.values()):
        tags.append("인턴")

    if "부문대표" in raw_role:
        return "담당디렉터", "부문대표", tags
    if "부대표" in raw_role:
        return "담당디렉터", "부대표", tags
    if "그룹장" in raw_role:
        return "그룹장", "그룹장", tags
    if "센터장" in raw_role:
        return "파트장/센터장", "센터장", tags
    if "파트장" in raw_role:
        return "파트장/센터장", "파트장", tags
    if raw_role == "담당":
        return "담당디렉터", "담당", tags
    if raw_role == "리더":
        return "시니어매니저", "리더", tags
    if "Sr.Manager" in raw_role or "Sr. Manager" in raw_role:
        return "시니어매니저", "Sr.매니저", tags
    return "매니저", "매니저", tags


def find_header(sheet) -> tuple[int, list[str]]:
    for row_idx in range(1, min(sheet.max_row, 20) + 1):
        headers = [clean(cell.value) for cell in sheet[row_idx]]
        if "성명" in headers and "회사 이메일" in headers and "그룹/파트/실/센터" in headers:
            return row_idx, headers
    raise ValueError("구성원 헤더 행을 찾지 못했습니다.")


def read_roster(source: Path, as_of: date) -> tuple[str, list[dict[str, Any]]]:
    workbook = load_workbook(source, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row, headers = find_header(sheet)
    people: list[dict[str, Any]] = []

    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        name = clean(row.get("성명"))
        if not name or clean(row.get("부문")) != "리얼에셋부문":
            continue
        leave_date = row.get("퇴사일 (계약만료일)") or row.get("퇴사일(계약만료일)")
        if isinstance(leave_date, datetime):
            leave_date = leave_date.date()
        if isinstance(leave_date, date) and leave_date < as_of:
            continue
        email = normalize_email(row.get("회사 이메일"))
        if not email:
            raise ValueError(f"회사 이메일이 없는 구성원: {name}")
        role, role_label, tags = classify_role(row)
        path = parse_org_path(clean(row.get("그룹/파트/실/센터")), clean(row.get("팀")))
        people.append(
            {
                "name": name,
                "email": email,
                "joinDate": to_iso_date(row.get("입사일")),
                "leaveDate": to_iso_date(leave_date),
                "title": clean(row.get("호칭(26년)")),
                "level": clean(row.get("직위(26년)")),
                "rawRole": clean(row.get("직책(26년)")),
                "role": role,
                "roleLabel": role_label,
                "tags": tags,
                "rawOrg": clean(row.get("그룹/파트/실/센터")),
                "rawTeam": clean(row.get("팀")),
                "concurrent": clean(row.get("겸직부서")),
                "path": path,
            }
        )

    emails = [person["email"] for person in people]
    if len(emails) != len(set(emails)):
        duplicates = sorted(email for email in set(emails) if emails.count(email) > 1)
        raise ValueError(f"중복 이메일이 있습니다: {duplicates}")
    return sheet.title, people


def path_scope_type(path: OrgPath) -> str:
    if path.part and path.part != "미지정" and not path.team:
        return "part"
    if not path.part and not path.team:
        return "group"
    return "team"


def build_token_lookup(people: list[dict[str, Any]]) -> dict[str, list[OrgPath]]:
    lookup: dict[str, set[OrgPath]] = defaultdict(set)
    for person in people:
        path: OrgPath = person["path"]
        raw_org = person["rawOrg"]
        raw_team = person["rawTeam"]
        lookup[raw_team].add(path)
        lookup[path.team].add(path)
        lookup[path.group].add(path.group_scope())
        if path.part and path.part != "미지정":
            lookup[raw_org].add(path.part_scope())
            raw_org_parts = [clean(value) for value in raw_org.split("/")]
            if len(raw_org_parts) > 1 and raw_org_parts[1]:
                lookup[raw_org_parts[1]].add(path.part_scope())
            lookup[path.part].add(path.part_scope())
            lookup[f"{path.group}{path.part}"].add(path.part_scope())
        else:
            lookup[raw_org].add(path.group_scope() if path.group.endswith("그룹") else path)
    return {token: sorted(paths, key=lambda value: value.key) for token, paths in lookup.items() if token}


def resolve_concurrent_path(token: str, lookup: dict[str, list[OrgPath]]) -> OrgPath | None:
    token = clean(token)
    if not token or token in IGNORED_CONCURRENT_TOKENS:
        return None
    if "TF" in token or "CFT" in token or token.endswith("실"):
        return None
    candidates = lookup.get(token, [])
    if token.endswith("센터"):
        team_candidates = [candidate for candidate in candidates if candidate.team]
        if len(team_candidates) == 1:
            return team_candidates[0]
    if token.endswith("파트") or token.endswith("그룹"):
        scoped_candidates = [candidate for candidate in candidates if not candidate.team]
        if len(scoped_candidates) == 1:
            return scoped_candidates[0]
    if len(candidates) == 1:
        return candidates[0]
    if token.endswith("센터") and token in GROUP_SECTION:
        group = normalize_group_name(token)
        return OrgPath(GROUP_SECTION[group], group, "미지정", group)
    return None


def member_from_person(
    person: dict[str, Any],
    *,
    role: str | None = None,
    role_label: str | None = None,
    extra_tags: Iterable[str] = (),
    assignment_kind: str,
) -> dict[str, Any]:
    tags = list(dict.fromkeys([*person["tags"], *extra_tags]))
    return {
        "role": role or person["role"],
        "roleLabel": role_label or person["roleLabel"],
        "rawName": person["name"],
        "name": person["name"],
        "email": person["email"],
        "tags": tags,
        "assignmentKind": assignment_kind,
    }


def concurrent_role(person: dict[str, Any], path: OrgPath) -> tuple[str, str]:
    scope = path_scope_type(path)
    if person["roleLabel"] in LEADER_LABELS:
        if scope == "group" and path.group.endswith("그룹"):
            return "그룹장", "그룹장"
        if scope == "part":
            return "파트장/센터장", "파트장"
        if path.group.endswith("센터"):
            return "파트장/센터장", "센터장"
    return person["role"], person["roleLabel"]


def build_assignments(people: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    token_lookup = build_token_lookup(people)
    assignments: list[dict[str, Any]] = []
    leader_scopes: list[tuple[OrgPath, dict[str, Any]]] = []
    unresolved_tokens: set[str] = set()
    for person in people:
        path: OrgPath = person["path"]
        label = person["roleLabel"]
        if label == "그룹장":
            leader_scopes.append((path.group_scope(), member_from_person(person, assignment_kind="primary-leader")))
        elif label == "파트장":
            leader_scopes.append((path.part_scope(), member_from_person(person, assignment_kind="primary-leader")))
        elif label == "센터장":
            leader_scopes.append((path, member_from_person(person, assignment_kind="primary-leader")))
        else:
            assignments.append({"path": path, "member": member_from_person(person, assignment_kind="primary")})

        for token in [clean(value) for value in person["concurrent"].split("/") if clean(value)]:
            target = resolve_concurrent_path(token, token_lookup)
            if target is None:
                if token not in IGNORED_CONCURRENT_TOKENS and "TF" not in token and "CFT" not in token and not token.endswith("실"):
                    unresolved_tokens.add(token)
                continue
            role, role_label = concurrent_role(person, target)
            member = member_from_person(
                person,
                role=role,
                role_label=role_label,
                extra_tags=("겸직",),
                assignment_kind="concurrent",
            )
            if role_label == "그룹장":
                leader_scopes.append((target.group_scope(), member))
            elif role_label == "파트장":
                leader_scopes.append((target.part_scope(), member))
            elif role_label == "센터장":
                leader_scopes.append((target, member))
            else:
                if not target.team:
                    fallback_part = target.part or "미지정"
                    fallback_team = fallback_part if fallback_part != "미지정" else target.group
                    target = OrgPath(target.section, target.group, fallback_part, fallback_team)
                assignments.append({"path": target, "member": member})

    team_paths = {entry["path"] for entry in assignments}

    for scope, member in leader_scopes:
        scope_type = path_scope_type(scope)
        if scope_type == "group":
            targets = [path for path in team_paths if path.section == scope.section and path.group == scope.group]
        elif scope_type == "part":
            targets = [
                path
                for path in team_paths
                if path.section == scope.section and path.group == scope.group and path.part == scope.part
            ]
        else:
            targets = [scope]
        if not targets:
            fallback_part = scope.part or "미지정"
            fallback_team = scope.team or (fallback_part if fallback_part != "미지정" else scope.group)
            targets = [OrgPath(scope.section, scope.group, fallback_part, fallback_team)]
        for target in targets:
            assignments.append({"path": target, "member": dict(member)})

    deduped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for entry in assignments:
        member = entry["member"]
        key = (member["email"], entry["path"].key, member["role"], member["roleLabel"])
        if key not in deduped:
            deduped[key] = entry
        else:
            existing_tags = deduped[key]["member"]["tags"]
            deduped[key]["member"]["tags"] = list(dict.fromkeys([*existing_tags, *member["tags"]]))

    role_rank = {role: index for index, role in enumerate(ROLE_ORDER)}
    result = sorted(
        deduped.values(),
        key=lambda entry: (
            SECTION_ORDER.index(entry["path"].section),
            GROUP_ORDER.get(entry["path"].section, []).index(entry["path"].group)
            if entry["path"].group in GROUP_ORDER.get(entry["path"].section, [])
            else 999,
            entry["path"].part,
            entry["path"].team,
            role_rank.get(entry["member"]["role"], 999),
            entry["member"]["name"],
        ),
    )
    return result, sorted(unresolved_tokens)


def unique_people(members: Iterable[dict[str, Any]]) -> int:
    return len({member.get("email") or member.get("name") for member in members if member.get("name")})


def build_payload(
    assignments: list[dict[str, Any]],
    tf_section: dict[str, Any] | None,
    source: Path,
    sheet_name: str,
    as_of: date,
) -> dict[str, Any]:
    by_path: dict[OrgPath, list[dict[str, Any]]] = defaultdict(list)
    for entry in assignments:
        by_path[entry["path"]].append(entry["member"])

    units: list[dict[str, Any]] = []
    for path, members in by_path.items():
        units.append(
            {
                "id": stable_id("unit", path.key, length=12),
                "section": path.section,
                "group": path.group,
                "part": path.part,
                "team": path.team,
                "displayName": path.team or path.part or path.group,
                "path": path.display_path,
                "assignmentCount": len(members),
                "uniquePeopleCount": unique_people(members),
                "members": members,
            }
        )

    sections: list[dict[str, Any]] = []
    for section_name in SECTION_ORDER:
        if section_name == "TFs":
            if tf_section:
                sections.append(tf_section)
            continue
        section_units = [unit for unit in units if unit["section"] == section_name]
        if not section_units:
            continue
        groups: list[dict[str, Any]] = []
        group_names = sorted(
            {unit["group"] for unit in section_units},
            key=lambda name: GROUP_ORDER.get(section_name, []).index(name)
            if name in GROUP_ORDER.get(section_name, [])
            else 999,
        )
        for group_name in group_names:
            group_units = [unit for unit in section_units if unit["group"] == group_name]
            parts: list[dict[str, Any]] = []
            for part_name in sorted({unit["part"] for unit in group_units}):
                part_units = [unit for unit in group_units if unit["part"] == part_name]
                part_members = [member for unit in part_units for member in unit["members"]]
                parts.append(
                    {
                        "name": part_name,
                        "assignmentCount": len(part_members),
                        "uniquePeopleCount": unique_people(part_members),
                        "teams": part_units,
                    }
                )
            group_members = [member for unit in group_units for member in unit["members"]]
            groups.append(
                {
                    "name": group_name,
                    "assignmentCount": len(group_members),
                    "uniquePeopleCount": unique_people(group_members),
                    "parts": parts,
                }
            )
        section_members = [
            member
            for group in groups
            for part in group["parts"]
            for team in part["teams"]
            for member in team["members"]
        ]
        sections.append(
            {
                "name": section_name,
                "assignmentCount": len(section_members),
                "uniquePeopleCount": unique_people(section_members),
                "groups": groups,
            }
        )

    all_members = [
        member
        for section in sections
        for group in section.get("groups", [])
        for part in group.get("parts", [])
        for team in part.get("teams", [])
        for member in team.get("members", [])
    ]
    return {
        "meta": {
            "sourceFile": source.name,
            "sheetName": sheet_name,
            "rosterAsOf": as_of.isoformat(),
            "generatedAt": datetime.now(KST).isoformat(timespec="seconds"),
            "organizationSource": "staff-roster-static",
        },
        "summary": {
            "uniquePeopleCount": unique_people(all_members),
            "sectionCount": len(sections),
            "groupCount": sum(len(section.get("groups", [])) for section in sections),
            "teamCount": sum(
                len(part.get("teams", []))
                for section in sections
                for group in section.get("groups", [])
                for part in group.get("parts", [])
            ),
        },
        "sections": sections,
        "units": units,
    }


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


class SupabaseRest:
    def __init__(self, base_url: str, key: str):
        self.base_url = base_url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}"}

    def request(
        self,
        method: str,
        path: str,
        payload: object | None = None,
        prefer: str = "",
    ) -> object | None:
        body = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers = {**self.headers, "Accept": "application/json"}
        if body is not None:
            headers["Content-Type"] = "application/json"
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(f"{self.base_url}/rest/v1/{path}", data=body, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                raw = response.read()
                return json.loads(raw.decode("utf-8")) if raw else None
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase {method} {path} failed ({error.code}): {detail}") from error

    def table(self, name: str) -> list[dict[str, Any]]:
        result = self.request("GET", f"{name}?select=*")
        if not isinstance(result, list):
            raise RuntimeError(f"{name} 조회 결과가 배열이 아닙니다.")
        return result

    def upsert(self, name: str, rows: list[dict[str, Any]], chunk_size: int = 150) -> None:
        # PostgREST requires every object in one bulk request to have the same keys.
        # Existing and newly created staff rows can differ in nullable audit columns.
        rows_by_shape: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            rows_by_shape[tuple(sorted(row))].append(row)
        conflict_column = "org_id" if name == "orgs" else "staff_id" if name == "staff" else "assignment_id"
        for shaped_rows in rows_by_shape.values():
            for index in range(0, len(shaped_rows), chunk_size):
                self.request(
                    "POST",
                    f"{name}?on_conflict={urllib.parse.quote(conflict_column)}",
                    shaped_rows[index : index + chunk_size],
                    "resolution=merge-duplicates,return=minimal",
                )

    def delete_ids(self, name: str, column: str, ids: list[str], chunk_size: int = 40) -> None:
        for index in range(0, len(ids), chunk_size):
            encoded = ",".join(f'"{value}"' for value in ids[index : index + chunk_size])
            self.request("DELETE", f"{name}?{column}=in.({urllib.parse.quote(encoded, safe='(),\"')})")


def normalize_section(value: str) -> str:
    value = clean(value).replace("&", "+")
    return "부문직속" if value in {"부분직속", "부문직속"} else value


def build_db_plan(
    people: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    db: dict[str, list[dict[str, Any]]],
    as_of: date,
) -> dict[str, Any]:
    org_by_id = {row["org_id"]: row for row in db["orgs"]}
    staff_by_id = {row["staff_id"]: row for row in db["staff"]}
    staff_by_email: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in db["staff"]:
        if normalize_email(row.get("email")):
            staff_by_email[normalize_email(row.get("email"))].append(row)

    matched_staff: dict[str, dict[str, Any]] = {}
    new_staff: list[dict[str, Any]] = []
    for person in people:
        matches = staff_by_email.get(person["email"], [])
        if len(matches) > 1:
            matches = sorted(matches, key=lambda row: (row.get("status") != "active", row["staff_id"]))
        if matches:
            matched_staff[person["email"]] = matches[0]
            continue
        staff_id = stable_id("staff_roster", person["email"], length=12)
        row = {
            "staff_id": staff_id,
            "employee_no": None,
            "name": person["name"],
            "eng_name": None,
            "email": person["email"],
            "title": person["title"] or None,
            "level": person["level"] or None,
            "position": person["roleLabel"],
            "org_id": None,
            "line_code": None,
            "line_label": None,
            "status": "active",
            "join_date": person["joinDate"] or None,
            "leave_date": person["leaveDate"] or None,
            "is_dual_role": False,
            "cohort": None,
            "notion_id": None,
            "source_system": "staff_roster_20260701",
            "metadata": {},
        }
        new_staff.append(row)
        matched_staff[person["email"]] = row
        staff_by_id[staff_id] = row

    existing_semantics: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    existing_group_scopes: dict[tuple[str, str], dict[str, Any]] = {}
    for org in db["orgs"]:
        key = (
            normalize_section(org.get("section")),
            clean(org.get("group_name")),
            clean(org.get("part_name")),
            clean(org.get("team_name")),
            clean(org.get("org_type")),
        )
        current = existing_semantics.get(key)
        if current is None or org["org_id"].startswith("org_section_"):
            existing_semantics[key] = org
        if not clean(org.get("part_name")) and not clean(org.get("team_name")) and clean(org.get("group_name")):
            scope_key = (normalize_section(org.get("section")), clean(org.get("group_name")))
            current_scope = existing_group_scopes.get(scope_key)
            if current_scope is None or org["org_id"].startswith("org_section_"):
                existing_group_scopes[scope_key] = org

    required_paths = sorted({entry["path"] for entry in assignments}, key=lambda path: path.key)
    new_orgs: list[dict[str, Any]] = []
    org_id_by_path: dict[OrgPath, str] = {}

    def ensure_org(
        section: str,
        group: str,
        part: str,
        team: str,
        org_type: str,
        parent_org_id: str | None,
    ) -> str:
        key = (normalize_section(section), group, part, team, org_type)
        existing = existing_semantics.get(key)
        if existing is None and org_type in {"group", "center"}:
            existing = existing_group_scopes.get((normalize_section(section), group))
        if existing:
            return existing["org_id"]
        org_id = stable_id("org_roster", *key, length=14)
        org_name = team or part or group or section
        row = {
            "org_id": org_id,
            "org_name": org_name,
            "org_type": org_type,
            "parent_org_id": parent_org_id,
            "section": section,
            "group_name": group or None,
            "part_name": part or None,
            "team_name": team or None,
            "org_path": " > ".join(value for value in (section, group, part, team) if value),
            "source_system": "staff_roster_20260701",
            "source_id": None,
            "is_active": True,
            "metadata": {"rosterManaged": True, "rosterAsOf": as_of.isoformat()},
        }
        existing_semantics[key] = row
        new_orgs.append(row)
        return org_id

    section_ids: dict[str, str] = {}
    group_ids: dict[tuple[str, str], str] = {}
    part_ids: dict[tuple[str, str, str], str] = {}
    for path in required_paths:
        section_id = section_ids.setdefault(
            path.section,
            ensure_org(path.section, "", "", "", "section", None),
        )
        group_key = (path.section, path.group)
        group_id = group_ids.setdefault(
            group_key,
            ensure_org(path.section, path.group, "", "", "group" if path.group.endswith("그룹") else "center", section_id),
        )
        part_key = (path.section, path.group, path.part)
        part_id = part_ids.setdefault(
            part_key,
            ensure_org(path.section, path.group, path.part, "", "part", group_id),
        )
        org_id_by_path[path] = ensure_org(path.section, path.group, path.part, path.team, "team", part_id)

    path_by_email: dict[str, list[OrgPath]] = defaultdict(list)
    for entry in assignments:
        path_by_email[entry["member"]["email"]].append(entry["path"])

    staff_updates: list[dict[str, Any]] = []
    person_by_email = {person["email"]: person for person in people}
    for email, staff in matched_staff.items():
        person = person_by_email[email]
        merged = dict(staff)
        metadata = dict(merged.get("metadata") or {})
        primary_path = person["path"]
        candidate_paths = path_by_email[email]
        if primary_path not in candidate_paths and candidate_paths:
            primary_path = candidate_paths[0]
        merged.update(
            {
                "title": person["title"] or merged.get("title"),
                "level": person["level"] or merged.get("level"),
                "position": person["roleLabel"],
                "org_id": org_id_by_path.get(primary_path) or (org_id_by_path[candidate_paths[0]] if candidate_paths else None),
                "join_date": person["joinDate"] or merged.get("join_date"),
                "leave_date": person["leaveDate"] or None,
                "is_dual_role": len({path.key for path in candidate_paths}) > 1,
                "metadata": {
                    **metadata,
                    "rosterName": person["name"],
                    "rosterAsOf": as_of.isoformat(),
                    "orgDashboardHidden": False,
                    "isIntern": "인턴" in person["tags"],
                    "currentMainOrgPath": primary_path.display_path,
                },
            }
        )
        for field in ("created_at", "updated_at"):
            merged.pop(field, None)
        staff_updates.append(merged)

    current_main_assignments: list[dict[str, Any]] = []
    tf_assignments: list[dict[str, Any]] = []
    for assignment in db["staff_org_assignments"]:
        org = org_by_id.get(assignment.get("org_id"), {})
        is_tf = normalize_section(org.get("section")) == "TFs" or re.search(
            r"TF|CFT", clean(org.get("group_name")), re.IGNORECASE
        )
        (tf_assignments if is_tf else current_main_assignments).append(assignment)

    source_staff_ids = {row["staff_id"] for row in matched_staff.values()}
    current_main_staff_ids = {row["staff_id"] for row in current_main_assignments}
    removed_staff_ids = sorted(current_main_staff_ids - source_staff_ids)
    removed_staff: list[dict[str, Any]] = []
    for staff_id in removed_staff_ids:
        existing = staff_by_id[staff_id]
        merged = dict(existing)
        metadata = dict(merged.get("metadata") or {})
        merged["org_id"] = None
        merged["metadata"] = {
            **metadata,
            "orgDashboardHidden": True,
            "orgDashboardRemovedAsOf": as_of.isoformat(),
        }
        for field in ("created_at", "updated_at"):
            merged.pop(field, None)
        removed_staff.append(merged)

    target_assignments: list[dict[str, Any]] = []
    primary_seen: set[str] = set()
    for entry in assignments:
        member = entry["member"]
        staff = matched_staff[member["email"]]
        org_id = org_id_by_path[entry["path"]]
        is_primary = member["email"] not in primary_seen and entry["path"] == person_by_email[member["email"]]["path"]
        if is_primary:
            primary_seen.add(member["email"])
        target_assignments.append(
            {
                "assignment_id": stable_id(
                    "org_roster_assignment",
                    staff["staff_id"],
                    org_id,
                    member["role"],
                    member["roleLabel"],
                    length=16,
                ),
                "staff_id": staff["staff_id"],
                "org_id": org_id,
                "role": member["role"],
                "raw_name": member["name"],
                "is_primary": is_primary,
                "is_dual_role": "겸직" in member["tags"],
                "source_system": "staff_roster_20260701",
                "metadata": {
                    "roleLabel": member["roleLabel"],
                    "tags": member["tags"],
                    "email": member["email"],
                    "rosterAsOf": as_of.isoformat(),
                    "assignmentKind": member["assignmentKind"],
                },
            }
        )

    source_emails = set(matched_staff)
    return {
        "new_orgs": new_orgs,
        "new_staff": new_staff,
        "staff_updates": staff_updates,
        "removed_staff": removed_staff,
        "delete_assignment_ids": [row["assignment_id"] for row in current_main_assignments],
        "target_assignments": target_assignments,
        "preserved_tf_assignments": tf_assignments,
        "removed_staff_summary": [
            {
                "staff_id": row["staff_id"],
                "name": row.get("name"),
                "email": row.get("email"),
                "reason": (
                    "legacy_duplicate_detached"
                    if normalize_email(row.get("email")) in source_emails
                    else "not_in_roster"
                ),
            }
            for row in removed_staff
        ],
        "new_staff_summary": [
            {"staff_id": row["staff_id"], "name": row.get("name"), "email": row.get("email")}
            for row in new_staff
        ],
    }


def write_dashboard(payload: dict[str, Any]) -> None:
    serialized = json.dumps(payload, ensure_ascii=False, indent=2)
    (DASHBOARD_DIR / "org-data.json").write_text(serialized + "\n", encoding="utf-8")
    (DASHBOARD_DIR / "org-data.js").write_text(
        f"window.ORG_DASHBOARD_DATA = {serialized};\n", encoding="utf-8"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="7월 구성원 명단을 조직 대시보드와 Supabase 조직 연결에 반영합니다.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--as-of", type=date.fromisoformat, default=DEFAULT_AS_OF)
    parser.add_argument("--write-dashboard", action="store_true")
    parser.add_argument("--apply-db", action="store_true")
    parser.add_argument("--env", type=Path, default=ROOT / ".env")
    parser.add_argument("--output-dir", type=Path, default=ROOT / "outputs" / "org_roster_sync_20260715")
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(args.source)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    existing_payload = json.loads((DASHBOARD_DIR / "org-data.json").read_text(encoding="utf-8-sig"))
    tf_section = next((section for section in existing_payload.get("sections", []) if section.get("name") == "TFs"), None)
    if tf_section is None:
        raise RuntimeError("보존할 TFs 섹션을 찾지 못했습니다.")

    sheet_name, people = read_roster(args.source, args.as_of)
    assignments, unresolved_tokens = build_assignments(people)
    payload = build_payload(assignments, tf_section, args.source, sheet_name, args.as_of)

    env = read_env(args.env)
    client = SupabaseRest(env["SUPABASE_URL"], env["SUPABASE_KEY"])
    db = {
        "staff": client.table("staff"),
        "orgs": client.table("orgs"),
        "staff_org_assignments": client.table("staff_org_assignments"),
    }
    plan = build_db_plan(people, assignments, db, args.as_of)

    backup_path = args.output_dir / "supabase_before.json"
    report_path = args.output_dir / "sync_report.json"
    backup_path.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    report = {
        "sourceFile": str(args.source),
        "asOf": args.as_of.isoformat(),
        "sourcePeople": len(people),
        "sourceInterns": sum("인턴" in person["tags"] for person in people),
        "dashboardMainAssignments": len(assignments),
        "dashboardMainUniquePeople": unique_people(entry["member"] for entry in assignments),
        "preservedTfAssignments": len(plan["preserved_tf_assignments"]),
        "newStaff": plan["new_staff_summary"],
        "removedFromMainOrg": plan["removed_staff_summary"],
        "newOrgRows": len(plan["new_orgs"]),
        "newOrgSummary": [
            {
                "org_id": row["org_id"],
                "org_type": row["org_type"],
                "org_path": row["org_path"],
            }
            for row in plan["new_orgs"]
        ],
        "oldMainAssignmentRows": len(plan["delete_assignment_ids"]),
        "newMainAssignmentRows": len(plan["target_assignments"]),
        "unresolvedConcurrentTokens": unresolved_tokens,
        "writeDashboard": args.write_dashboard,
        "applyDb": args.apply_db,
    }

    if args.write_dashboard:
        write_dashboard(payload)

    if args.apply_db:
        # Permission probe: targets no rows and leaves the database unchanged.
        client.request(
            "PATCH",
            "staff?staff_id=eq.__codex_org_sync_permission_probe__",
            {"updated_at": datetime.now(KST).isoformat()},
            "return=minimal",
        )
        client.upsert("orgs", plan["new_orgs"])
        client.upsert("staff", [*plan["staff_updates"], *plan["removed_staff"]])
        # Publish the replacement rows before removing the old main-org links.
        # A failed insert therefore cannot leave the dashboard with an empty roster.
        client.upsert("staff_org_assignments", plan["target_assignments"])
        client.delete_ids("staff_org_assignments", "assignment_id", plan["delete_assignment_ids"])

        verified = {
            "staff": client.table("staff"),
            "orgs": client.table("orgs"),
            "staff_org_assignments": client.table("staff_org_assignments"),
        }
        verified_org_by_id = {row["org_id"]: row for row in verified["orgs"]}
        verified_main = [
            row
            for row in verified["staff_org_assignments"]
            if normalize_section(verified_org_by_id.get(row.get("org_id"), {}).get("section")) != "TFs"
        ]
        verified_tf = [
            row
            for row in verified["staff_org_assignments"]
            if normalize_section(verified_org_by_id.get(row.get("org_id"), {}).get("section")) == "TFs"
        ]
        if len(verified_tf) != len(plan["preserved_tf_assignments"]):
            raise RuntimeError("TF 연결 건수가 변경되었습니다.")
        if len(verified_main) != len(plan["target_assignments"]):
            raise RuntimeError(
                f"일반 조직 연결 검증 실패: expected={len(plan['target_assignments'])}, actual={len(verified_main)}"
            )
        if len(verified["staff"]) < len(db["staff"]):
            raise RuntimeError("staff 행 수가 감소했습니다.")
        report["verified"] = {
            "staffRowsBefore": len(db["staff"]),
            "staffRowsAfter": len(verified["staff"]),
            "mainAssignmentsAfter": len(verified_main),
            "tfAssignmentsAfter": len(verified_tf),
        }

    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
