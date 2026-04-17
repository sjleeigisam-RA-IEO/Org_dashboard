import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


BASE = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = BASE / "CCMM빌딩 좌석배치도_260403_업로드용.xlsx"
JSON_PATH = BASE / "ccmm-seat-layout-data.json"
JS_PATH = BASE / "ccmm-seat-layout-data.js"
UPLOAD_PATH = BASE / "ccmm_layout_google_sheet_upload.xlsx"

OUR_SEAT_RE = re.compile(r"^C([1-9]|[1-4][0-9]|5[0-5])$")
ALL_SEAT_RE = re.compile(r"^[CD](\d{1,3})$")
OFFICE_LABEL_RE = re.compile(r"(부대표실|그룹장실|파트장실|센터장실)")


def classify_shape(label: str) -> str:
    text = (label or "").strip()
    if not text:
        return "common"
    if "승강기홀" in text:
        return "core"
    if "회의실" in text:
        return "room"
    if OFFICE_LABEL_RE.search(text):
        return "office"
    if "휴게" in text:
        return "common"
    if any(word in text for word in ["창고", "OA", "서버실"]):
        return "support"
    return "common"


def seat_sort_key(code: str):
    match = re.match(r"^([CD])(\d+)$", code or "")
    if not match:
        return ("Z", 999)
    return (match.group(1), int(match.group(2)))


def parse_seat_sheet(ws):
    seats = {}
    names_by_seat = {}
    seat_by_name = {}
    all_seat_cells = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            seat_label = value.strip()
            if not ALL_SEAT_RE.fullmatch(seat_label):
                continue
            all_seat_cells[seat_label] = cell.coordinate
            if not OUR_SEAT_RE.fullmatch(seat_label):
                continue
            person_value = ws.cell(row=cell.row + 1, column=cell.column).value
            person_name = "" if person_value is None else str(person_value).strip()
            seats[seat_label] = {
                "label": seat_label,
                "coord": cell.coordinate,
                "x": cell.column - 1,
                "y": cell.row - 1,
            }
            names_by_seat[seat_label] = person_name
            if person_name:
                seat_by_name[person_name] = seat_label
    return seats, names_by_seat, seat_by_name, all_seat_cells


def parse_shapes(ws):
    shapes = []
    seen_anchors = set()

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        label = ws.cell(min_row, min_col).value
        if not label:
            continue
        label = str(label).strip()
        shape_type = classify_shape(label)
        if shape_type == "core":
            continue
        shapes.append(
            {
                "shapeId": f"CCMM11F-shape-{len(shapes)+1:03d}",
                "floorCode": "CCMM11F",
                "shapeType": shape_type,
                "label": label,
                "x": min_col - 1,
                "y": min_row - 1,
                "w": max_col - min_col + 1,
                "h": max_row - min_row + 1,
                "fill": None,
                "source": str(merged_range),
            }
        )
        seen_anchors.add((min_row, min_col))

    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in seen_anchors:
                continue
            if not isinstance(cell.value, str):
                continue
            label = cell.value.strip()
            if not OFFICE_LABEL_RE.search(label):
                continue
            shapes.append(
                {
                    "shapeId": f"CCMM11F-shape-{len(shapes)+1:03d}",
                    "floorCode": "CCMM11F",
                    "shapeType": "office",
                    "label": label,
                    "x": cell.column - 1,
                    "y": cell.row - 1,
                    "w": 2,
                    "h": 4,
                    "fill": None,
                    "source": cell.coordinate,
                }
            )
    return shapes


def parse_core_areas(ws):
    cores = []
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        label = ws.cell(min_row, min_col).value
        if not label:
            continue
        label = str(label).strip()
        if "승강기홀" not in label:
            continue
        cores.append(
            {
                "label": label,
                "x": min_col - 1,
                "y": min_row - 1,
                "w": max_col - min_col + 1,
                "h": max_row - min_row + 1,
            }
        )
    return cores


def build_floor_outline(shapes, seats):
    min_x = min([shape["x"] for shape in shapes] + [seat["x"] for seat in seats.values()]) - 1
    min_y = min([shape["y"] for shape in shapes] + [seat["y"] for seat in seats.values()]) - 1
    max_x = max([shape["x"] + shape["w"] for shape in shapes] + [seat["x"] + 1 for seat in seats.values()]) + 1
    max_y = max([shape["y"] + shape["h"] for shape in shapes] + [seat["y"] + 2 for seat in seats.values()]) + 1
    return [[min_x, min_y], [max_x, min_y], [max_x, max_y], [min_x, max_y]]


def build_payload():
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    current_ws = workbook.worksheets[0]
    plan_ws = workbook.worksheets[1]

    seats, current_names, current_by_name, _ = parse_seat_sheet(current_ws)
    _, plan_names, _, _ = parse_seat_sheet(plan_ws)
    shapes = parse_shapes(plan_ws)
    core_areas = parse_core_areas(plan_ws)
    outline_points = build_floor_outline(shapes, seats)

    seat_defs = []
    current_assignments = {}
    plan_assignments = {}
    upload_rows = []

    for seat_label, seat in sorted(seats.items(), key=lambda item: seat_sort_key(item[0])):
        seat_code = f"CCMM11F-{seat_label}"
        current_person = current_names.get(seat_label, "")
        plan_person = plan_names.get(seat_label, "")
        origin_seat = current_by_name.get(plan_person, "") if plan_person else ""
        origin_floor = "CCMM11F" if origin_seat else ""
        is_moved = "Y" if plan_person and origin_seat and origin_seat != seat_label else "N"

        seat_defs.append(
            {
                "seatCode": seat_code,
                "seatLabel": seat_label,
                "x": seat["x"],
                "y": seat["y"],
                "w": 1,
                "h": 1,
                "sourceCell": seat["coord"],
            }
        )
        current_assignments[seat_code] = {"personName": current_person, "deptName": ""}
        plan_assignments[seat_code] = {"personName": plan_person, "deptName": ""}

        upload_rows.append(
            {
                "seat_code": seat_code,
                "floor_code": "CCMM11F",
                "seat_label": seat_label,
                "person_name": plan_person,
                "team_org_id_seat": "",
                "origin_floor_code": origin_floor,
                "origin_seat_code": f"CCMM11F-{origin_seat}" if origin_seat else "",
                "is_moved": is_moved,
                "is_external_division": "N",
                "note": "",
            }
        )

    payload = {
        "meta": {
            "sourceWorkbook": WORKBOOK_PATH.name,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "floorOrder": ["CCMM11F"],
        },
        "floors": [
            {
                "floorCode": "CCMM11F",
                "displayName": "CCMM11층",
                "outlinePoints": outline_points,
                "coreAreas": core_areas,
                "shapes": shapes,
                "seatDefs": seat_defs,
                "scenarios": {
                    "current": {"assignments": current_assignments},
                    "plan": {"assignments": plan_assignments},
                },
            }
        ],
    }
    return payload, upload_rows


def write_js(payload):
    JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    JS_PATH.write_text(
        "window.SEAT_LAYOUT_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def write_upload(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ccmm_layout_latest"
    headers = [
        "seat_code",
        "floor_code",
        "seat_label",
        "person_name",
        "team_org_id_seat",
        "origin_floor_code",
        "origin_seat_code",
        "is_moved",
        "is_external_division",
        "note",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    widths = {"A": 18, "B": 12, "C": 10, "D": 14, "E": 24, "F": 14, "G": 18, "H": 10, "I": 18, "J": 24}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    guide = workbook.create_sheet("guide")
    guide_rows = [
        ["목적", "CCMM 11층 자리배치 최신 상태를 구글시트에서 관리하기 위한 업로드용 시트입니다."],
        ["기준", "한 행 = 한 좌석 (우리 부문 C1~C55만 포함)"],
        ["필수 수정 컬럼", "person_name, team_org_id_seat, origin_floor_code, origin_seat_code, is_moved, note"],
        ["seat_code", "불변 좌석 코드. 수정하지 않습니다."],
        ["floor_code", "층 코드. CCMM11F로 고정합니다."],
        ["team_org_id_seat", "조직 레이어 표시용 조직 코드입니다."],
        ["origin_seat_code", "이동자면 원래 좌석 코드를 입력합니다."],
        ["is_moved", "이동자면 Y, 아니면 N"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 74

    workbook.save(UPLOAD_PATH)


def main():
    payload, rows = build_payload()
    write_js(payload)
    write_upload(rows)
    print(JSON_PATH)
    print(JS_PATH)
    print(UPLOAD_PATH)


if __name__ == "__main__":
    main()
