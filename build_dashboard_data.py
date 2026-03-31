from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = BASE_DIR.parent / "260325_리얼에셋부문_조직개편_인력배치_v1.4_RA.xlsx"
OUTPUT_JS = BASE_DIR / "org-data.js"
OUTPUT_JSON = BASE_DIR / "org-data.json"

ROLE_BY_COLUMN = {
    4: "그룹장",
    5: "파트장/센터장",
    6: "담당디렉터",
    7: "담당디렉터",
    8: "담당디렉터",
    9: "담당디렉터",
    10: "시니어매니저",
    11: "시니어매니저",
    12: "시니어매니저",
    13: "시니어매니저",
    14: "시니어매니저",
    15: "매니저",
    16: "매니저",
    17: "매니저",
    18: "매니저",
    19: "매니저",
    20: "매니저",
    21: "매니저",
    22: "매니저",
    23: "매니저",
    24: "매니저",
    25: "매니저",
}

ROLE_ORDER = ["디렉터", "그룹장", "파트장/센터장", "담당디렉터", "시니어매니저", "매니저"]
EXCLUDED_SECTIONS = {"크레딧"}
EXCLUDED_UNIT_NAMES = {"기타"}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " / ").strip()


def clean_org_label(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"\s*/\s*\(.*?\)", "", text)
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"\s*/\s*$", "", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()


def normalize_name(raw: str) -> str:
    name = raw.split("/")[0].strip()
    name = re.sub(r"\(겸\)|\(대\)", "", name).strip()
    return name


def detect_tags(raw: str) -> list[str]:
    tags: list[str] = []
    if "겸" in raw:
        tags.append("겸직")
    if "(대)" in raw:
        tags.append("대행")
    if "외부영입" in raw:
        tags.append("외부영입")
    return tags


def resolve_role(display_name: str, col_idx: int, raw: str) -> str:
    if display_name == "IOTA CFT":
        return "디렉터"
    return ROLE_BY_COLUMN[col_idx]


def extract_assignments(row, display_name: str) -> list[dict]:
    assignments: list[dict] = []
    for col_idx in ROLE_BY_COLUMN:
        raw = clean_text(row[col_idx - 1].value)
        if not raw:
            continue
        assignments.append(
            {
                "role": resolve_role(display_name, col_idx, raw),
                "rawName": raw,
                "name": normalize_name(raw),
                "tags": detect_tags(raw),
            }
        )
    return assignments


def is_section_header(label_a: str, label_b: str, label_c: str, assignments: list[dict]) -> bool:
    return bool(label_a and not label_b and not label_c and not assignments)


def build_hierarchy(units: list[dict]) -> list[dict]:
    sections: dict[str, dict] = {}
    for unit in units:
        section = sections.setdefault(unit["section"], {"name": unit["section"], "groups": {}})
        group = section["groups"].setdefault(
            unit["group"],
            {"name": unit["group"], "parts": {}, "members": []},
        )
        group["members"].extend(unit["members"])

        part_name = unit["part"] or "미지정"
        part = group["parts"].setdefault(
            part_name,
            {"name": part_name, "teams": [], "members": []},
        )
        part["members"].extend(unit["members"])
        part["teams"].append(unit)

    ordered_sections: list[dict] = []
    for section in sections.values():
        groups = []
        for group in section["groups"].values():
            parts = []
            for part in group["parts"].values():
                parts.append(
                    {
                        "name": part["name"],
                        "assignmentCount": len(part["members"]),
                        "uniquePeopleCount": len({m["name"] for m in part["members"] if m["name"]}),
                        "teams": part["teams"],
                    }
                )
            groups.append(
                {
                    "name": group["name"],
                    "assignmentCount": len(group["members"]),
                    "uniquePeopleCount": len({m["name"] for m in group["members"] if m["name"]}),
                    "parts": parts,
                }
            )
        ordered_sections.append(
            {
                "name": section["name"],
                "assignmentCount": sum(group["assignmentCount"] for group in groups),
                "uniquePeopleCount": len(
                    {
                        member["name"]
                        for group in groups
                        for part in group["parts"]
                        for team in part["teams"]
                        for member in team["members"]
                        if member["name"]
                    }
                ),
                "groups": groups,
            }
        )
    return ordered_sections


def build_role_seat_counts(units: list[dict]) -> list[dict]:
    seats_by_role: dict[str, set[str]] = {role: set() for role in ROLE_ORDER}
    group_leader_by_group: dict[tuple[str, str], str] = {}

    for unit in units:
        group_key = (unit["section"], unit["group"])
        for member in unit["members"]:
            if member["role"] == "그룹장" and group_key not in group_leader_by_group:
                group_leader_by_group[group_key] = member["name"]

    for unit in units:
        group_key = (unit["section"], unit["group"])
        group_leader_name = group_leader_by_group.get(group_key)
        for member in unit["members"]:
            role = member["role"]
            if role == "디렉터":
                seat_key = f"{unit['path']}|{member['name']}|{role}"
            elif role == "그룹장":
                seat_key = f"{unit['section']}|{unit['group']}|{role}"
            elif role == "파트장/센터장":
                if member["name"] and member["name"] == group_leader_name:
                    continue
                part_scope = unit["part"] or unit["displayName"]
                seat_key = f"{unit['section']}|{unit['group']}|{part_scope}|{role}"
            else:
                seat_key = f"{unit['path']}|{member['name']}|{role}"
            seats_by_role[role].add(seat_key)

    return [{"role": role, "count": len(seats_by_role[role])} for role in ROLE_ORDER]


def summarize(units: list[dict], sections: list[dict]) -> dict:
    members = [member for unit in units for member in unit["members"]]
    unique_people = sorted({member["name"] for member in members if member["name"]})
    part_names = {
        unit["part"]
        for unit in units
        if unit["part"] and unit["part"] != "미지정"
    }
    section_counter = [
        {
            "name": section["name"],
            "uniquePeopleCount": section["uniquePeopleCount"],
        }
        for section in sections
    ]
    return {
        "uniquePeopleCount": len(unique_people),
        "sectionCount": len(sections),
        "groupCount": len({unit["group"] for unit in units}),
        "partCount": len(part_names),
        "teamCount": len(units),
        "roleSeatCounts": build_role_seat_counts(units),
        "sectionCounts": section_counter,
    }


def parse_workbook(source_path: Path) -> dict:
    workbook = load_workbook(source_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    units: list[dict] = []
    current_section = ""
    current_group = ""
    current_part = ""

    for row_idx in range(4, sheet.max_row + 1):
        row = sheet[row_idx]
        label_a = clean_org_label(row[0].value)
        label_b = clean_org_label(row[1].value)
        label_c = clean_org_label(row[2].value)
        prospective_display_name = label_c or label_b or label_a
        assignments = extract_assignments(row, prospective_display_name)

        if not (label_a or label_b or label_c or assignments):
            if row_idx > 60:
                break
            continue

        if "붉은색 이탤릭체" in label_a:
            break

        if is_section_header(label_a, label_b, label_c, assignments):
            current_section = label_a
            current_group = ""
            current_part = ""
            continue

        if current_section in EXCLUDED_SECTIONS:
            continue

        if not assignments:
            continue

        if label_a:
            current_group = label_a
            current_part = label_b or ""
        elif label_b:
            current_part = label_b

        if label_b and label_a:
            current_part = label_b

        part_name = label_b or current_part or ""
        team_name = label_c
        display_name = team_name or part_name or current_group

        if display_name in EXCLUDED_UNIT_NAMES or label_a in EXCLUDED_UNIT_NAMES:
            continue

        units.append(
            {
                "id": f"unit-{len(units) + 1}",
                "section": current_section or "미분류",
                "group": current_group or "미분류",
                "part": part_name,
                "team": team_name,
                "displayName": display_name,
                "path": " > ".join(
                    value for value in [current_section, current_group, part_name, team_name] if value
                ),
                "assignmentCount": len(assignments),
                "uniquePeopleCount": len({member["name"] for member in assignments if member["name"]}),
                "members": assignments,
            }
        )

    sections = build_hierarchy(units)
    summary = summarize(units, sections)

    return {
        "meta": {
            "sourceFile": source_path.name,
            "sheetName": sheet.title,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
        },
        "summary": summary,
        "sections": sections,
        "units": units,
    }


def main() -> None:
    data = parse_workbook(DEFAULT_SOURCE)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(payload, encoding="utf-8")
    OUTPUT_JS.write_text(f"window.ORG_DASHBOARD_DATA = {payload};\n", encoding="utf-8")
    print(f"Created {OUTPUT_JSON}")
    print(f"Created {OUTPUT_JS}")


if __name__ == "__main__":
    main()
