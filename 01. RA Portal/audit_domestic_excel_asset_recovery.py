from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
WORKLIST_CSV = OUTPUT_DIR / "ra_insight_db_asset_name_cleanup_worklist.csv"
EXCEL_GLOB = "Deal Board DB_v20240812*base.xlsx"
MATCHES_CSV = OUTPUT_DIR / "asset_name_recovery_domestic_excel_matches.csv"
ADOPTION_CSV = OUTPUT_DIR / "asset_name_recovery_adoption_candidates.csv"
AUDIT_MD = OUTPUT_DIR / "asset_name_recovery_domestic_excel_audit.md"


def cell_text(row, idx: int) -> str:
    if idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    text = str(value).strip().replace("\xa0", " ")
    if text.endswith(".0"):
        return text[:-2]
    return text


def md_cell(value: object) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "<br>").replace("\n", " ")


def split_pipe(value: str) -> list[str]:
    return [item.strip() for item in value.split("|") if item.strip()]


def load_domestic_asset_rows() -> list[dict[str, str]]:
    excel_files = list(OUTPUT_DIR.glob(EXCEL_GLOB))
    if not excel_files:
        raise FileNotFoundError(f"No workbook matched {EXCEL_GLOB}")

    workbook = load_workbook(excel_files[0], read_only=True, data_only=True)
    # Sheet: 운용자산_국내_완료_update_반기
    worksheet = workbook.worksheets[5]

    rows = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(
            {
                "pnu": cell_text(raw, 0),
                "x_code": cell_text(raw, 1),
                "y_code": cell_text(raw, 2),
                "modified_address": cell_text(raw, 3),
                "fund_code": cell_text(raw, 4),
                "short_name": cell_text(raw, 5),
                "fund_name": cell_text(raw, 6),
                "investment_asset_type": cell_text(raw, 7),
                "seq": cell_text(raw, 8),
                "excel_asset_name": cell_text(raw, 9),
                "asset_category": cell_text(raw, 11),
                "city": cell_text(raw, 12),
                "district": cell_text(raw, 13),
                "address": cell_text(raw, 14),
                "operation_status": cell_text(raw, 21),
                "domestic_overseas": cell_text(raw, 25),
                "fund_class": cell_text(raw, 26),
            }
        )
    return rows


def unique_values(rows: list[dict[str, str]], key: str) -> list[str]:
    values = []
    for row in rows:
        value = row.get(key, "")
        if value and value not in values:
            values.append(value)
    return values


def normalize_name(value: str) -> str:
    return "".join(str(value).lower().split())


def build_name_index(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    index: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        name = normalize_name(row.get("excel_asset_name", ""))
        if name:
            index[name].append(row)
    return index


def find_name_hits(
    candidate: str,
    exact_index: dict[str, list[dict[str, str]]],
    domestic_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    normalized = normalize_name(candidate)
    if not normalized:
        return []
    if normalized in exact_index:
        return exact_index[normalized]
    hits = []
    for row in domestic_rows:
        excel_name = normalize_name(row.get("excel_asset_name", ""))
        if not excel_name:
            continue
        if normalized in excel_name or excel_name in normalized:
            hits.append(row)
    return hits


def main() -> None:
    domestic_rows = load_domestic_asset_rows()
    by_fund: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in domestic_rows:
        by_fund[row["fund_code"]].append(row)
    by_name = build_name_index(domestic_rows)

    with WORKLIST_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        worklist_rows = list(csv.DictReader(handle))

    problem_rows = [row for row in worklist_rows if row["asset_name_action"] != "keep"]

    matched_assets = []
    unmatched_assets = []
    match_rows = []
    adoption_rows = []

    for row in problem_rows:
        hits = []
        for fund_id in split_pipe(row["linked_fund_ids"]):
            hits.extend(by_fund.get(fund_id, []))

        if not hits:
            unmatched_assets.append(row)
            for candidate in split_pipe(row["proposed_asset_name"]) or [
                row["proposed_asset_name"]
            ]:
                name_hits = find_name_hits(candidate, by_name, domestic_rows)
                if name_hits:
                    for hit in name_hits:
                        adoption_rows.append(
                            {
                                "asset_id": row["asset_id"],
                                "current_asset_name": row["current_asset_name"],
                                "accepted_asset_name": candidate,
                                "asset_name_action": row["asset_name_action"],
                                "adoption_source": "worklist_proposed_name_matched_domestic_excel_by_name",
                                "linked_fund_ids": row["linked_fund_ids"],
                                **hit,
                            }
                        )
                else:
                    adoption_rows.append(
                        {
                            "asset_id": row["asset_id"],
                            "current_asset_name": row["current_asset_name"],
                            "accepted_asset_name": candidate,
                            "asset_name_action": row["asset_name_action"],
                            "adoption_source": "worklist_proposed_name_unmatched",
                            "linked_fund_ids": row["linked_fund_ids"],
                            "pnu": "",
                            "x_code": "",
                            "y_code": "",
                            "modified_address": "",
                            "fund_code": "",
                            "short_name": "",
                            "fund_name": "",
                            "investment_asset_type": "",
                            "seq": "",
                            "excel_asset_name": "",
                            "asset_category": "",
                            "city": "",
                            "district": "",
                            "address": "",
                            "operation_status": "",
                            "domestic_overseas": "",
                            "fund_class": "",
                        }
                    )
            continue

        matched_assets.append((row, hits))
        for hit in hits:
            record = {
                "asset_id": row["asset_id"],
                "current_asset_name": row["current_asset_name"],
                "asset_name_action": row["asset_name_action"],
                "proposed_asset_name": row["proposed_asset_name"],
                "linked_fund_ids": row["linked_fund_ids"],
                **hit,
            }
            match_rows.append(record)
            adoption_rows.append(
                {
                    "asset_id": row["asset_id"],
                    "current_asset_name": row["current_asset_name"],
                    "accepted_asset_name": hit["excel_asset_name"],
                    "asset_name_action": row["asset_name_action"],
                    "adoption_source": "domestic_excel_fund_code_match",
                    "linked_fund_ids": row["linked_fund_ids"],
                    **hit,
                }
            )

    match_fields = [
        "asset_id",
        "current_asset_name",
        "asset_name_action",
        "proposed_asset_name",
        "linked_fund_ids",
        "pnu",
        "x_code",
        "y_code",
        "modified_address",
        "fund_code",
        "short_name",
        "fund_name",
        "investment_asset_type",
        "seq",
        "excel_asset_name",
        "asset_category",
        "city",
        "district",
        "address",
        "operation_status",
        "domestic_overseas",
        "fund_class",
    ]
    with MATCHES_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=match_fields)
        writer.writeheader()
        writer.writerows(match_rows)

    adoption_fields = [
        "asset_id",
        "current_asset_name",
        "accepted_asset_name",
        "asset_name_action",
        "adoption_source",
        "linked_fund_ids",
        "pnu",
        "x_code",
        "y_code",
        "modified_address",
        "fund_code",
        "short_name",
        "fund_name",
        "investment_asset_type",
        "seq",
        "excel_asset_name",
        "asset_category",
        "city",
        "district",
        "address",
        "operation_status",
        "domestic_overseas",
        "fund_class",
    ]
    with ADOPTION_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=adoption_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(adoption_rows)

    action_counter = Counter(row["asset_name_action"] for row, _ in matched_assets)
    action_row_counter = Counter()
    for row, hits in matched_assets:
        action_row_counter[row["asset_name_action"]] += len(hits)

    excel_file = list(OUTPUT_DIR.glob(EXCEL_GLOB))[0].name
    lines: list[str] = []
    lines.append("# 국내 운용자산 원장 기반 자산명 복원 검토")
    lines.append("")
    lines.append(f"- 원장 파일: `{excel_file}`")
    lines.append("- 사용 시트: `운용자산_국내_완료_update_반기`")
    lines.append("- 기준 작업표: `ra_insight_db_asset_name_cleanup_worklist.csv`")
    lines.append(f"- 원장 행 수: {len(domestic_rows)}행, 고유 펀드코드: {len(by_fund)}개")
    lines.append(f"- 검토 대상: `asset_name_action != keep` {len(problem_rows)}개 asset_id")
    lines.append(
        f"- 원장 매칭: {len(matched_assets)}개 asset_id / {len(problem_rows)}개, 원장 상세행 {len(match_rows)}행"
    )
    lines.append(f"- 원장 미매칭: {len(unmatched_assets)}개 asset_id")
    lines.append(
        f"- 채택 후보 파일: `{ADOPTION_CSV.name}` ({len(adoption_rows)}행, 미매칭 대상은 제안 후보 채택)"
    )
    lines.append(
        f"- 원장 매칭 상세행 중 PNU 보유: {sum(1 for row in match_rows if row.get('pnu'))}/{len(match_rows)}행"
    )
    lines.append(
        f"- 원장 매칭 상세행 중 X/Y 좌표 보유: {sum(1 for row in match_rows if row.get('x_code') and row.get('y_code'))}/{len(match_rows)}행"
    )
    lines.append(
        f"- 채택 후보 전체 중 주소 보유: {sum(1 for row in adoption_rows if row.get('address') or row.get('modified_address'))}/{len(adoption_rows)}행"
    )
    lines.append("")
    lines.append("## 매칭 요약")
    lines.append("")
    lines.append("| 구분 | asset_id 수 | 원장 상세행 | 처리 의미 |")
    lines.append("|---|---:|---:|---|")
    descriptions = {
        "update_asset_name": "단일 자산명 교체 확인",
        "split_or_drawer_list": "펀드 선택 시 자산 리스트 drawer 필요",
        "manual_review": "이번 원장에서는 직접 매칭 없음",
    }
    for action in ["update_asset_name", "split_or_drawer_list", "manual_review"]:
        lines.append(
            f"| `{action}` | {action_counter[action]} | {action_row_counter[action]} | {descriptions[action]} |"
        )

    lines.append("")
    lines.append("## 핵심 발견")
    lines.append("")
    lines.append(
        "- 국내 원장에는 펀드코드별 `자산(건물)명`과 `전체주소`가 분리되어 있어, 국내 실물/특별자산의 대량 보정 원천으로 쓸 수 있습니다."
    )
    lines.append(
        "- `이지스일반사모투자신탁435호(태양광)`은 단일 자산명이 아니라 원장 기준 46개 태양광 자산입니다. 현재 DB 후보 43개보다 원장이 3개 더 많습니다: `미션`, `소망`, `청수`."
    )
    lines.append(
        "- `이지스일반사모투자신탁392호(태양광)`도 원장 기준 26개 태양광 자산입니다. 현재 DB 후보 23개보다 원장이 3개 더 많습니다: `임원`, `누리`, `한려`."
    )
    lines.append(
        "- 404호 물류센터, 리테일55호, 코어리테일302호처럼 여러 기초자산이 있는 펀드는 원장과 DB 후보가 서로 맞습니다. 이들은 자산명을 하나로 바꾸지 말고 drawer/list 관계로 유지하는 편이 맞습니다."
    )
    lines.append(
        "- 원장 미매칭 18개는 해외/재간접/증권성/리츠 SMA 또는 국내 원장에 없는 PFV/SPC·프로젝트성 코드가 섞여 있습니다. 이번 국내 완료 운용자산 시트만으로는 보강하기 어렵습니다."
    )
    lines.append(
        "- 미매칭 18개도 이번 기준에서는 `proposed_asset_name`을 채택 후보로 펼쳤습니다. 다만 주소/PNU/좌표는 원장명 검색으로 찾힌 일부를 제외하면 비어 있습니다."
    )

    lines.append("")
    lines.append("## 원장 매칭 asset_id 상세")
    lines.append("")
    lines.append("| action | asset_id | 현재 자산명 | 연결 펀드 | 원장 자산 수 | 주소 수 | PNU 수 | 좌표 수 | 원장 자산명 |")
    lines.append("|---|---|---|---|---:|---:|---:|---:|---|")
    for row, hits in matched_assets:
        names = unique_values(hits, "excel_asset_name")
        addresses = unique_values(hits, "address")
        pnus = unique_values(hits, "pnu")
        coords = {
            (hit.get("x_code", ""), hit.get("y_code", ""))
            for hit in hits
            if hit.get("x_code") and hit.get("y_code")
        }
        preview = " | ".join(names[:10]) + (" ..." if len(names) > 10 else "")
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{row['asset_name_action']}`",
                    f"`{row['asset_id']}`",
                    md_cell(row["current_asset_name"]),
                    md_cell(row["linked_fund_ids"]),
                    str(len(names)),
                    str(len(addresses)),
                    str(len(pnus)),
                    str(len(coords)),
                    md_cell(preview),
                ]
            )
            + " |"
        )

    lines.append("")
    lines.append("## 태양광 펀드 상세 차이")
    for fund_id in ["112584", "112504"]:
        linked = [
            row
            for row in problem_rows
            if fund_id in split_pipe(row["linked_fund_ids"])
        ]
        if not linked:
            continue
        candidates = split_pipe(linked[0]["proposed_asset_name"])
        excel_asset_names = [
            row["excel_asset_name"] for row in by_fund[fund_id] if row["excel_asset_name"]
        ]
        missing = [name for name in excel_asset_names if name not in candidates]
        lines.append("")
        lines.append(f"### 펀드코드 {fund_id} / {linked[0]['current_asset_name']}")
        lines.append("")
        lines.append(f"- 원장 자산 수: {len(excel_asset_names)}개")
        lines.append(f"- 기존 DB 후보 수: {len(candidates)}개")
        lines.append(
            f"- DB 후보에 없고 원장에 있는 자산: {', '.join(missing) if missing else '없음'}"
        )
        lines.append("")
        lines.append("| 순번 | 자산명 | 주소 | PNU | X | Y |")
        lines.append("|---:|---|---|---|---:|---:|")
        for row in by_fund[fund_id]:
            lines.append(
                f"| {md_cell(row['seq'])} | {md_cell(row['excel_asset_name'])} | {md_cell(row['address'])} | {md_cell(row['pnu'])} | {md_cell(row['x_code'])} | {md_cell(row['y_code'])} |"
            )

    lines.append("")
    lines.append("## 원장 미매칭 대상")
    lines.append("")
    lines.append("| action | asset_id | 현재 자산명 | 연결 펀드 | 제안/후보 |")
    lines.append("|---|---|---|---|---|")
    for row in unmatched_assets:
        lines.append(
            "| "
            + " | ".join(
                [
                    f"`{row['asset_name_action']}`",
                    f"`{row['asset_id']}`",
                    md_cell(row["current_asset_name"]),
                    md_cell(row["linked_fund_ids"]),
                    md_cell(row["proposed_asset_name"]),
                ]
            )
            + " |"
        )

    lines.append("")
    lines.append("## 채택 후보 주소/좌표 상태")
    lines.append("")
    source_counts = Counter(row["adoption_source"] for row in adoption_rows)
    lines.append("| source | 행 수 | 주소 보유 | PNU 보유 | 좌표 보유 |")
    lines.append("|---|---:|---:|---:|---:|")
    for source, count in source_counts.items():
        source_rows = [row for row in adoption_rows if row["adoption_source"] == source]
        lines.append(
            f"| `{source}` | {count} | "
            f"{sum(1 for row in source_rows if row.get('address') or row.get('modified_address'))} | "
            f"{sum(1 for row in source_rows if row.get('pnu'))} | "
            f"{sum(1 for row in source_rows if row.get('x_code') and row.get('y_code'))} |"
        )

    AUDIT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(
        {
            "problem_assets": len(problem_rows),
            "matched_assets": len(matched_assets),
            "matched_rows": len(match_rows),
            "unmatched_assets": len(unmatched_assets),
            "matched_actions": dict(action_counter),
            "report": str(AUDIT_MD),
            "matches_csv": str(MATCHES_CSV),
            "adoption_csv": str(ADOPTION_CSV),
        }
    )


if __name__ == "__main__":
    main()
